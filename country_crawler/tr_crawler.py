# -*- coding: utf-8 -*-
"""土耳其 TR 爬虫：ODMD 零售销量（type=36 XLSX，2004至今全品牌）"""
import re
import io
import logging
from datetime import datetime, date

import requests
from openpyxl import load_workbook

from base_crawler import BaseCrawler, DB_CONFIG

logger = logging.getLogger('tr_crawler')

TR_BASE = 'https://www.odmd.org.tr/web_2837_1'
TR_TYPE36 = '/neuralnetwork.aspx?type=36'
TR_LIST = '/sortial.aspx?type=36&target=categorial1&primary_id=&detail=single&sp_table=&sp_primary=&sp_table_extra=&openfrom=sortial&linkpos={n}&language_id=1&search_fields=&search_values='
TR_DOWNLOAD = '/wf_docudownload.aspx?primary_id={id}&type=36&target=categorial1&detail=single&sp_table=&sp_primary=&sp_table_extra=&openfrom=sortial&downloadfirst=yes'

TR_MONTH_MAP = {
    'ocak': 1, 'şubat': 2, 'subat': 2, 'mart': 3, 'nisan': 4, 'mayıs': 5, 'mayis': 5,
    'haziran': 6, 'temmuz': 7, 'ağustos': 8, 'agustos': 8, 'eylül': 9, 'eylul': 9,
    'ekim': 10, 'kasım': 11, 'kasim': 11, 'aralık': 12, 'aralik': 12,
}

# 品牌级直接取 TOPLAM 总列（OTOMOBİL+HAFİF TİCARİ 合计），vehicle_type='passenger'
# 若需拆分：c2-4=OTOMOBİL(YERLİ/İTHAL/TOPLAM) c5-7=HAFİF TİCARİ c8-10=TOPLAM
# TOPLAM_COL=10（1-based），品牌列=1


class TrCrawler(BaseCrawler):
    def __init__(self, source_name='odmd_tr_retail_sales', country_code='TR'):
        super().__init__(source_name, country_code)
        self.session = requests.Session()
        from base_crawler import UA_LIST
        self.session.headers.update({
            'User-Agent': UA_LIST[0],
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
        })
        self._brand_id_cache = {}
        self._report_index = None
        self._dl_prefix = None

    # ---------- 报告索引 ----------
    def _build_report_index(self):
        """遍历 linkpos=1..17，解析报告行：{(y,m): {'id':.., 'title':..}}（仅单月报告）
        YTD 累计报告标题含 '(Ocak-...)'，跳过。"""
        # 先建立 ASP.NET_SessionId（三步走第1步）
        try:
            self.retry_request(self.session.get, TR_BASE + TR_TYPE36, timeout=60)
        except Exception:
            pass
        index = {}
        for n in range(1, 18):
            url = TR_BASE + TR_LIST.format(n=n)
            try:
                r = self.retry_request(self.session.get, url, timeout=60)
            except Exception:
                continue
            if not r:
                continue
            r.encoding = 'utf-8'
            html = r.text
            # 报告行 <tr class="tr_r_rep">... sRep(id) ... td_reports_name ...</tr>
            rows = re.findall(r'<tr[^>]*tr_r_rep[^>]*>(.*?)</tr>', html, re.S)
            if not rows:
                rows = re.findall(r'<tr[^>]*>(.*?)</tr>', html, re.S)
            for row in rows:
                m_id = re.search(r"sRep\((\d+)\)", row)
                m_t = re.search(r'td_reports_name[^>]*>\s*([^<]+?)\s*<', row)
                if not m_id or not m_t:
                    continue
                rid = m_id.group(1)
                title = m_t.group(1).strip()
                # 匹配单月报告，兼容格式变体：
                #   "2026 Temmuz Perakende"
                #   "2015 Yılı Ocak Perakende" / "2017 Yılı Kasım Perakende"
                #   "2016 Yılı (Şubat) Perakende" / "2016 Yılı (Ocak) Perakende"
                # 跳过年度报告（"2009 Yılı Perakende"、"2007 Yılı Toplam"）与YTD累计（"(Ocak-...)"）
                mm = re.match(
                    r'^(\d{4})\s+(?:Yılı\s+)?\(?\s*([A-Za-zÇĞİÖŞÜçğıöşü]+)\s*\)?\s+Perakende',
                    title)
                if not mm:
                    continue
                year = int(mm.group(1))
                month = TR_MONTH_MAP.get(mm.group(2).lower())
                if month:
                    index[(year, month)] = {'id': rid, 'title': title}
            # 分页结束判断：无 next 或已到末尾
            if 'linkpos=' + str(n + 1) not in html and n >= 17:
                break
        self._report_index = index
        return index

    def get_report_id(self, year, month):
        if self._report_index is None:
            self._build_report_index()
        return self._report_index.get((year, month))

    def latest_available_month(self):
        if self._report_index is None:
            self._build_report_index()
        if not self._report_index:
            return None
        return max(self._report_index.keys())

    # ---------- 下载 ----------
    def _download_report(self, rid):
        url = TR_BASE + TR_DOWNLOAD.format(id=rid)
        try:
            r = self.retry_request(self.session.get, url, timeout=120,
                                   headers={'Referer': TR_BASE + TR_LIST.format(n=1)})
        except Exception:
            return None
        if not r or r.status_code != 200:
            return None
        if len(r.content) < 1000:
            return None  # 433 空壳页
        return r.content

    # ---------- 解析 ----------
    def _parse_pdf(self, content, year, month):
        """旧版 PDF 报告解析：单页，'MARKA' 表头后的每行 = 品牌 + 数字序列。
        每行最后一个数字 = TOPLAM 总销量（品牌总销量，带千分位点号如 1.017=1017）。
        品牌=行首字母序列；遇 TOPLAM: 合计行跳过。"""
        records = []
        try:
            import pdfplumber
            with pdfplumber.open(io.BytesIO(content)) as pdf:
                txt = '\n'.join((p.extract_text() or '') for p in pdf.pages)
        except Exception:
            return []
        lines = txt.split('\n')
        in_table = False
        for ln in lines:
            s = ln.strip()
            if not s:
                continue
            # 进入表区：表头行含 MARKA
            if s.upper().startswith('MARKA') and 'OTOMOBİL' not in s.upper():
                in_table = True
                continue
            if not in_table:
                continue
            # 结束：脚注行（ODD / 来源），TOPLAM 合计行
            if 'ODD' in s.upper() and 'Derneği' in s:
                break
            # 品牌行：字母开头 + 数字
            m = re.match(r'^([A-Za-zÇĞİÖŞÜçğıöşü][A-Za-zÇĞİÖŞÜçğıöşü\s\.\-\']*?)\s+([\d\.\s]+)$', s)
            if not m:
                continue
            brand = re.sub(r'\s+', ' ', m.group(1)).strip()
            # 表头续行（OTOMOBİL/HAFİF TİCARİ/YERLİ/İTHAL/TOPLAM 等无数字）
            num_part = m.group(2)
            if not re.search(r'\d', num_part):
                continue
            nums = re.findall(r'\d[\d\.]*', num_part)
            if not nums:
                continue
            # 每行最后一个数字 = TOPLAM 总销量
            qty = int(nums[-1].replace('.', ''))
            bu = brand.upper()
            if bu in ('MARKA', 'TOPLAM', 'TOTAL', 'GENEL TOPLAM') or bu.startswith('TOPLAM') or bu.startswith('GENEL'):
                continue
            if qty <= 0:
                continue
            records.append({
                'country_code': 'TR',
                'source_month': date(year, month, 1),
                'brand_name_raw': brand,
                'brand_id': self.get_brand_id(brand),
                'model_name': None,
                'vehicle_type': 'passenger',
                'energy_type': None,
                'segment': None,
                'raw_unit': 'units',
                'sales_volume_raw': qty,
                'sales_volume_normalized': qty,
                'revision_no': 1,
                'is_latest': True,
                'pub_date': None,
                'crawl_time': datetime.now(),
                'data_source': self.source_name,
                'notes': 'ODMD perakende satışlar (toplam PDF, otomobil + hafif ticari)',
            })
        return records

    def parse_xlsx(self, content, year, month):
        """品牌=col1，销量=TOPLAM总列 col10（1-based）
        兼容两种格式：XLSX（PK magic，openpyxl）与旧 XLS（OLE magic，pandas+xlrd）。"""
        records = []
        rows = []
        if content[:4] == b'PK\x03\x04':
            # XLSX
            try:
                wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
            except Exception:
                return []
            ws = wb.active
            for row in ws.iter_rows():
                vals = [cell.value for cell in row]
                rows.append(vals)
        elif content[:8] == b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1':
            # 旧 XLS（OLE）
            try:
                import pandas as pd
                df = pd.read_excel(io.BytesIO(content), header=None, engine='xlrd')
                rows = df.astype(object).values.tolist()
            except Exception:
                return []
        elif content[:4] == b'%PDF':
            # 旧版（2004-2011年左右）PDF 报告：单页文本表
            # 每行 = MARKA + 数字序列，最后一个数字 = TOPLAM（品牌总销量，含小数千分位点号）
            return self._parse_pdf(content, year, month)
        else:
            return []
        for vals in rows:
            c1 = vals[0] if len(vals) > 0 else None
            if c1 is None:
                continue
            brand = str(c1).strip()
            if not brand or brand.lower() in ('nan', 'none'):
                continue
            bu = brand.upper()
            if bu in ('MARKA', 'TOPLAM', 'TOTAL', 'GENEL TOPLAM') or bu.startswith('TOPLAM') or bu.startswith('GENEL'):
                continue
            # TOPLAM 总列 = 第10列（索引9）
            qty = vals[9] if len(vals) > 9 else None
            if qty is None:
                continue
            try:
                qty = int(float(qty))
            except (ValueError, TypeError):
                continue
            if qty <= 0:
                continue
            records.append({
                'country_code': 'TR',
                'source_month': date(year, month, 1),
                'brand_name_raw': brand,
                'brand_id': self.get_brand_id(brand),
                'model_name': None,
                'vehicle_type': 'passenger',
                'energy_type': None,
                'segment': None,
                'raw_unit': 'units',
                'sales_volume_raw': qty,
                'sales_volume_normalized': qty,
                'revision_no': 1,
                'is_latest': True,
                'pub_date': None,
                'crawl_time': datetime.now(),
                'data_source': self.source_name,
                'notes': 'ODMD perakende satışlar (toplam, otomobil + hafif ticari)',
            })
        return records

    def get_brand_id(self, brand_raw):
        key = brand_raw.strip().upper()
        if key in self._brand_id_cache:
            return self._brand_id_cache[key]
        conn, cur = self.get_connection()
        bid = None
        cur.execute("""SELECT id FROM brand_name_mapping
                       WHERE UPPER(canonical_name)=%s OR UPPER(brand_name_cn)=%s
                       ORDER BY CASE WHEN status='active' THEN 0 ELSE 1 END, id LIMIT 1""",
                    (key, key))
        row = cur.fetchone()
        if row:
            bid = row['id'] if isinstance(row, dict) else row[0]
        else:
            cur.execute("""SELECT brand_id FROM brand_name_variant
                           WHERE UPPER(variant_name)=%s LIMIT 1""", (key,))
            row = cur.fetchone()
            if row:
                bid = row['brand_id'] if isinstance(row, dict) else row[0]
        self._brand_id_cache[key] = bid
        return bid

    def save_sales(self, record):
        super().save_sales(record)

    def crawl_month(self, year, month):
        rep = self.get_report_id(year, month)
        if not rep:
            return {'records': 0, 'reason': 'no report'}
        content = self._download_report(rep['id'])
        if not content:
            return {'records': 0, 'reason': 'download fail'}
        records = self.parse_xlsx(content, year, month)
        for rec in records:
            self.save_sales(rec)
        return {'records': len(records)}

    def _get_db_max_month(self):
        conn, cur = self.get_connection()
        cur.execute("SELECT MAX(source_month) AS m FROM market_sales_monthly WHERE country_code='TR'")
        row = cur.fetchone()
        m = row['m'] if isinstance(row, dict) else row[0]
        if isinstance(m, datetime):
            return m.date()
        return m

    def crawl_incremental(self):
        latest = self.latest_available_month()
        if not latest:
            return 0
        ly, lm = latest
        max_m = self._get_db_max_month()
        if max_m and date(ly, lm, 1) <= max_m:
            return 0
        return self.crawl_month(ly, lm).get('records', 0)

    def crawl_range(self, start_year, start_month, end_year, end_month):
        total = 0
        y, m = start_year, start_month
        while (y, m) <= (end_year, end_month):
            res = self.crawl_month(y, m)
            total += res.get('records', 0)
            m += 1
            if m > 12:
                m = 1
                y += 1
        return total


if __name__ == '__main__':
    import sys
    import io as _io
    sys.stdout = _io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    c = TrCrawler()
    print('latest:', c.latest_available_month())
    print('incremental:', c.crawl_incremental())
