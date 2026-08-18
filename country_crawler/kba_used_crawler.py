# -*- coding: utf-8 -*-
"""德国 DE 二手车爬虫: KBA FZ 9.1 乘用车所有权转移注册(Besitzumschreibungen)
口径: 二手车所有权转移(二手交易/转手), 非新车。月度 XLSX 直链。
"""
import re
import io
import time
import requests
from datetime import date, datetime
from base_crawler import BaseCrawler, DB_CONFIG, UA_LIST

KBA_USE_BASE = 'https://www.kba.de/SharedDocs/Downloads/DE/Statistik/Fahrzeuge/FZ9'
KBA_USE_DL = f'{KBA_USE_BASE}/fz9_{{y:04d}}{{m:02d}}.xlsx?__blob=publicationFile&v=2'

SKIP = ('TOTAL', 'INSGESAMT', 'DAVON', 'NACHHER', 'DARUNTER')


class KbaUsedCrawler(BaseCrawler):
    def __init__(self):
        super().__init__('kba_de_used_car_transfers_monthly', 'DE')
        self.session = requests.Session()
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36',
        })
        self._brand_cache = {}

    def _fetch(self, url, timeout=60):
        for attempt in range(3):
            try:
                r = self.session.get(url, timeout=timeout)
                if r.status_code == 200 and len(r.content) > 1000:
                    return r.content
            except Exception as e:
                print(f'_fetch err {url}: {e}')
            time.sleep(2)
        return None

    def parse_fz91(self, xlsx_content):
        """解析 FZ 9.1 sheet, 返回 [(brand, qty)]。
        sheet名含 'FZ 9.1'; 行9子表头后数据, 品牌=B列, 当月=C列。
        """
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(xlsx_content), data_only=True, read_only=True)
        target = None
        for ws in wb.worksheets:
            if 'FZ 9.1' in ws.title:
                target = ws
                break
        if target is None:
            return []
        rows = list(target.iter_rows(values_only=True))
        # 找表头行: 行9 (0-based 8) 子表头, 行8 (0-based 7) 主表头含 'Marke'
        # 数据从行10 (0-based 9) 起
        data_start = None
        for i, row in enumerate(rows[:15]):
            if row and any(str(c).strip() == 'Anzahl' for c in row if c is not None):
                data_start = i + 1
                break
        if data_start is None:
            return []
        brands = []
        for row in rows[data_start:]:
            if not row:
                continue
            b = row[1]  # B列 品牌
            if b is None:
                continue
            b = str(b).strip()
            if not b:
                continue
            bu = b.upper()
            if bu in SKIP or bu.startswith(('TOTAL', 'INSGESAMT', 'DAVON', 'NACHHER', 'DARUNTER')):
                continue
            qty = row[2]  # C列 当月Anzahl
            if qty is None:
                continue
            try:
                q = int(qty)
            except (ValueError, TypeError):
                continue
            if q <= 0:
                continue
            brands.append((b, q))
        return brands

    def get_brand_id(self, brand_raw):
        b = str(brand_raw).strip().upper()
        if b in self._brand_cache:
            return self._brand_cache[b]
        bid = None
        conn, cur = self.get_connection()
        cur.execute("SELECT id FROM brand_name_mapping WHERE UPPER(canonical_name)=%s OR UPPER(brand_name_cn)=%s ORDER BY CASE WHEN status='active' THEN 0 ELSE 1 END, id LIMIT 1", (b, b))
        row = cur.fetchone()
        if row:
            bid = row['id'] if isinstance(row, dict) else row[0]
        if bid is None:
            cur.execute("SELECT brand_id FROM brand_name_variant WHERE UPPER(variant_name)=%s ORDER BY id LIMIT 1", (b,))
            row = cur.fetchone()
            if row:
                bid = row['brand_id'] if isinstance(row, dict) else row[0]
        self._brand_cache[b] = bid
        return bid

    def save_used(self, brand, qty, y, m):
        """写入 market_used_vehicle_monthly"""
        conn, cur = self.get_connection()
        bid = self.get_brand_id(brand)
        cur.execute("""
            INSERT INTO market_used_vehicle_monthly
                (country_code, source_month, brand_name_raw, brand_id, vehicle_type,
                 used_volume, used_import_volume, data_source, notes, crawl_time, created_at)
            VALUES (%s, %s, %s, %s, 'passenger', %s, NULL, %s, %s, %s, NOW())
            ON CONFLICT (country_code, source_month, brand_name_raw, vehicle_type, data_source)
            DO UPDATE SET used_volume=EXCLUDED.used_volume, brand_id=EXCLUDED.brand_id, crawl_time=NOW()
        """, ('DE', date(y, m, 1), brand, bid, qty,
              self.source_name, 'KBA FZ9.1 Besitzumschreibungen Pkw (used car ownership transfers)', datetime.now()))

    def crawl_month(self, year, month):
        url = KBA_USE_DL.format(y=year, m=month)
        content = self._fetch(url)
        if not content:
            return {'records': 0}
        brands = self.parse_fz91(content)
        conn, cur = self.get_connection()
        for b, q in brands:
            self.save_used(b, q, year, month)
        conn.commit()
        return {'records': len(brands)}

    def _get_db_max_month(self):
        conn, cur = self.get_connection()
        cur.execute("SELECT MAX(source_month) AS m FROM market_used_vehicle_monthly WHERE country_code='DE' AND data_source=%s", (self.source_name,))
        row = cur.fetchone()
        m = row['m'] if isinstance(row, dict) else row[0]
        return m.date() if hasattr(m, 'date') else m

    def crawl_incremental(self):
        """探测最新月(当月或上月), 若>库MAX则爬"""
        today = date.today()
        for offset in (0, 1):
            y, m = today.year, today.month - offset
            if m <= 0:
                m += 12
                y -= 1
            content = self._fetch(KBA_USE_DL.format(y=y, m=m))
            if not content:
                continue
            brands = self.parse_fz91(content)
            if not brands:
                continue
            max_m = self._get_db_max_month()
            sm = date(y, m, 1)
            if max_m is not None and sm <= max_m:
                return 0
            conn, cur = self.get_connection()
            for b, q in brands:
                self.save_used(b, q, y, m)
            conn.commit()
            return len(brands)
        return 0

    def crawl_range(self, y1, m1, y2, m2):
        conn, cur = self.get_connection()
        total = 0
        yy, mm = y1, m1
        while (yy, mm) <= (y2, m2):
            url = KBA_USE_DL.format(y=yy, m=mm)
            content = self._fetch(url)
            if content:
                brands = self.parse_fz91(content)
                for b, q in brands:
                    self.save_used(b, q, yy, mm)
                conn.commit()
                total += len(brands)
                print(f'{yy}-{mm:02d}: {len(brands)} brands')
            else:
                print(f'{yy}-{mm:02d}: no data')
            mm += 1
            if mm > 12:
                mm = 1
                yy += 1
            time.sleep(1)
        return total


def main():
    import argparse
    ap = argparse.ArgumentParser()
    ap.add_argument('--incremental', action='store_true')
    ap.add_argument('--ym', type=str, default='', help='单月 2026-07')
    ap.add_argument('--y1', type=int, default=2024)
    ap.add_argument('--m1', type=int, default=1)
    ap.add_argument('--y2', type=int, default=2026)
    ap.add_argument('--m2', type=int, default=7)
    args = ap.parse_args()

    c = KbaUsedCrawler()
    if args.ym:
        y, m = map(int, args.ym.split('-'))
        print(c.crawl_month(y, m))
    elif args.incremental:
        n = c.crawl_incremental()
        print(f'DE used incremental saved: {n}')
    else:
        n = c.crawl_range(args.y1, args.m1, args.y2, args.m2)
        print(f'DE used range done: {n}')


if __name__ == '__main__':
    main()
