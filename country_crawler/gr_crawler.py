# -*- coding: utf-8 -*-
import sys, io, os, re, random, time
from datetime import datetime, date

import requests
import psycopg2
from psycopg2.extras import RealDictCursor

from base_crawler import BaseCrawler, DB_CONFIG

GR_BASE_URL = 'https://seaa.gr'
GR_ARCHIVE_URL = 'https://seaa.gr/passenger-car-registrations-comparisons/'
GR_UA = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36'

SKIP_BRANDS = ('TOTAL', 'TAXI', 'TAXIS', 'OTHER', 'OTHERS')


class GrCrawler(BaseCrawler):
    def __init__(self):
        super().__init__(source_name='seaa', country_code='GR')
        self.session = requests.Session()
        self.session.headers['User-Agent'] = GR_UA
        self._file_index = None
        self._brand_id_cache = {}

    # ---------- discovery ----------
    def _build_file_index(self):
        if self._file_index is not None:
            return self._file_index
        r = requests.get(GR_ARCHIVE_URL, headers={'User-Agent': GR_UA}, timeout=30)
        r.raise_for_status()
        idx = {}
        for m in re.finditer(r'href="([^"]+\.xlsx)"', r.text):
            url = m.group(1)
            mm = re.search(r'/(\d{4})-(\d{1,2})-comp[^/]*\.xlsx', url)
            if mm:
                y, mo = int(mm.group(1)), int(mm.group(2))
                if 1 <= mo <= 12:
                    idx[(y, mo)] = url
        self._file_index = idx
        return idx

    def discover_download_url(self, year, month):
        idx = self._build_file_index()
        return idx.get((year, month))

    def latest_available_month(self):
        idx = self._build_file_index()
        keys = [k for k in idx if k[0] >= 2000]
        if not keys:
            return None
        y, m = max(keys)
        return (y, m)

    def download_excel(self, url):
        r = self.retry_request(self.session.get, url, timeout=60)
        r.raise_for_status()
        return r.content

    # ---------- parse ----------
    def parse_excel(self, content):
        import io as _io
        import openpyxl
        wb = openpyxl.load_workbook(_io.BytesIO(content), data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        # locate header row (col2 contains 'Brand' or 'Make')
        header_idx = None
        for i, row in enumerate(rows[:10]):
            if len(row) >= 2 and row[1] is not None:
                h = str(row[1]).upper()
                if 'BRAND' in h or 'MAKE' in h:
                    header_idx = i
                    break
        if header_idx is None:
            return []
        records = []
        for row in rows[header_idx + 1:]:
            if len(row) < 3:
                continue
            brand = row[1]
            qty = row[2]
            if brand is None:
                continue
            brand = str(brand).strip()
            if not brand or brand.upper() in SKIP_BRANDS:
                continue
            if brand.upper().startswith('TOTAL'):
                continue
            try:
                qty = int(float(qty))
            except (TypeError, ValueError):
                continue
            if qty <= 0:
                continue
            records.append((brand, qty))
        return records

    # ---------- db helpers ----------
    def get_brand_id(self, brand_name_raw):
        if brand_name_raw in self._brand_id_cache:
            return self._brand_id_cache[brand_name_raw]
        conn, cur = self.get_connection()
        lookup = brand_name_raw.upper()
        cur.execute("SELECT id FROM brand_name_mapping WHERE (UPPER(canonical_name)=%s OR UPPER(brand_name_cn)=%s) ORDER BY CASE WHEN status='active' THEN 0 ELSE 1 END, id LIMIT 1", (lookup, lookup))
        row = cur.fetchone()
        if row and 'id' in row:
            bid = row['id']
        elif row and 'brand_id' in row:
            bid = row['brand_id']
        else:
            cur.execute("SELECT v.brand_id FROM brand_name_variant v WHERE UPPER(v.variant_name)=%s LIMIT 1", (lookup,))
            row = cur.fetchone()
            bid = (row.get('brand_id') if row else None)
        self._brand_id_cache[brand_name_raw] = bid
        return bid

    def save_sales(self, record):
        record['brand_id'] = self.get_brand_id(record['brand_name_raw'])
        super().save_sales(record)

    def _get_db_max_month(self):
        conn, cur = self.get_connection()
        cur.execute("SELECT MAX(source_month) AS m FROM market_sales_monthly WHERE country_code='GR' AND data_source='seaa'")
        row = cur.fetchone()
        return row['m'] if row and row['m'] else None

    def crawl_month(self, year, month):
        url = self.discover_download_url(year, month)
        if not url:
            return {'records': 0, 'msg': f'no file for {year}-{month:02d}'}
        content = self.download_excel(url)
        records = self.parse_excel(content)
        n = 0
        for brand, qty in records:
            rec = {
                'country_code': 'GR', 'source_month': date(year, month, 1),
                'brand_name_raw': brand, 'model_name': None,
                'vehicle_type': 'passenger', 'energy_type': None,
                'segment': None, 'raw_unit': 'units',
                'sales_volume_raw': qty, 'sales_volume_normalized': qty,
                'revision_no': 1, 'is_latest': True, 'pub_date': None,
                'crawl_time': datetime.now(), 'data_source': 'seaa',
                'notes': 'SEAA passenger car registrations by brand (comp)',
            }
            self.save_sales(rec)
            n += 1
        return {'records': n, 'msg': f'{year}-{month:02d}: {n} brands'}

    def crawl_incremental(self):
        latest = self.latest_available_month()
        if not latest:
            return 0
        max_d = self._get_db_max_month()
        latest_d = date(latest[0], latest[1], 1)
        if max_d and latest_d <= max_d:
            return 0
        result = self.crawl_month(latest[0], latest[1])
        return result['records']

    def crawl_range(self, sy, sm, ey, em):
        out = {}
        y, m = sy, sm
        while (y, m) <= (ey, em):
            out[f'{y}-{m:02d}'] = self.crawl_month(y, m)
            m += 1
            if m > 12:
                m = 1; y += 1
        return out


if __name__ == '__main__':
    c = GrCrawler()
    print('index months:', len(c._build_file_index()))
    print('latest:', c.latest_available_month())
