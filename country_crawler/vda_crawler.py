# -*- coding: utf-8 -*-
"""德国 DE 产量爬虫: VDA (vda.de) 月度 Pkw 产量
数据源: VDA Monatszahlen Excel (2016-01 起, 行12=Produktion im Inland)
+ sitemap 发现月度新闻稿作增量校验。
产量口径: Pkw 乘用车国内生产, Quelle: VDA/KBA。
"""
import re
import io
import time
import requests
from datetime import date, datetime
from openpyxl import load_workbook
from base_crawler import BaseCrawler, DB_CONFIG, UA_LIST

VDA_ARCHIV_URL = 'https://www.vda.de/dam/jcr:794adcf4-cfb6-4a6e-989f-753e9020f72d/Daten%20Internetarchiv%20Ausgabedatei_d.xlsx'
VDA_SITEMAP = 'https://www.vda.de/de/sitemap.xml'
VDA_MONAT = 'https://www.vda.de/de/aktuelles/zahlen-und-daten/monatszahlen'

GER_MONTHS = {'Januar': 1, 'Februar': 2, 'März': 3, 'April': 4, 'Mai': 5, 'Juni': 6,
              'Juli': 7, 'August': 8, 'September': 9, 'Oktober': 10, 'November': 11, 'Dezember': 12}


class VdaCrawler(BaseCrawler):
    def __init__(self):
        super().__init__('vda_production', 'DE')
        self.session = requests.Session()
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36',
            'Accept-Language': 'de,en;q=0.9',
        })

    def _fetch(self, url, timeout=60):
        for attempt in range(3):
            try:
                r = self.session.get(url, timeout=timeout)
                if r.status_code == 200:
                    return r
            except Exception as e:
                print(f'_fetch {url} err: {e}')
            time.sleep(2)
        return None

    def download_archive(self):
        """下载历史全量 Excel (2016-01 起), 返回 bytes"""
        r = self._fetch(VDA_ARCHIV_URL, timeout=120)
        return r.content if r else None

    def parse_archive(self, content):
        """解析 Excel: 行12 'Produktion im Inland' = Pkw 月度产量。
        行2=年份, 行3=月份, 每12列一年。返回 {(y, m): prod_volume}"""
        wb = load_workbook(io.BytesIO(content), data_only=True)
        ws = wb[wb.sheetnames[0]]
        # 找行12
        prod_row = None
        for r in range(1, 20):
            v = ws.cell(r, 1).value
            if v and 'Produktion im Inland' in str(v):
                prod_row = r
                break
        if prod_row is None:
            return {}
        # 遍历列2..max, 年份行2/月份行3
        result = {}
        for c in range(2, ws.max_column + 1):
            year = ws.cell(2, c).value
            month = ws.cell(3, c).value
            val = ws.cell(prod_row, c).value
            if year is None or month is None:
                continue
            # 年份可能是 2016 / 2017 跨列
            try:
                y = int(year)
            except (TypeError, ValueError):
                continue
            m = GER_MONTHS.get(str(month).strip())
            if m is None:
                continue
            try:
                v = int(val)
            except (TypeError, ValueError):
                continue
            result[(y, m)] = v
        return result

    def _insert(self, y, m, vol):
        conn, cur = self.get_connection()
        cur.execute("""
            INSERT INTO market_production_monthly
                (country_code, source_month, vehicle_type, energy_type, production_volume, data_source, notes, crawl_time)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            ON CONFLICT (country_code, source_month, vehicle_type, energy_type, data_source)
            DO UPDATE SET production_volume = EXCLUDED.production_volume,
                          crawl_time = EXCLUDED.crawl_time
        """, ('DE', date(y, m, 1), 'passenger', None, vol,
              'vda_production', 'VDA Monatszahlen Produktion im Inland (Pkw)', datetime.now()))
        conn.commit()

    def crawl_archive(self):
        """全量入库 Excel 历史"""
        content = self.download_archive()
        if not content:
            return 0
        data = self.parse_archive(content)
        n = 0
        for (y, m), vol in sorted(data.items()):
            self._insert(y, m, vol)
            n += 1
        print(f'VDA archive saved {n} months ({(min(data) if data else None)}~{(max(data) if data else None)})')
        return {'records': n, 'months': len(data)}

    def crawl_latest(self):
        """从统计页/最新新闻稿抓最新月产量 (增量)"""
        # 用统计页 monatszahlen (含最新月表)
        r = self._fetch(VDA_MONAT)
        if not r:
            return 0
        html = r.text
        # 找 'Produktion in Deutschland' 行的当月数字
        # 表结构: 行头 'Produktion in Deutschland' + 数字(当月/同比/累计/同比)
        m = re.search(r'Produktion in Deutschland\s*</t[dh]>\s*<t[dh][^>]*>\s*([\d\.]+)', html)
        saved = 0
        if m:
            vol = int(m.group(1).replace('.', ''))
            # 确定月份: 页面标题含 'März 2026' 等
            mm = re.search(r'([A-ZÄÖÜa-zäöü]+)\s*(\d{4})', re.sub(r'<[^>]+>', ' ', html)[:3000])
            # 页面有 'Stand' 或表头 'Juli 2026'
            saved = 0
            # 简化: 由调度器传目标年月调用 _insert
        return {'records': saved}

    def crawl_incremental(self):
        """增量: 统计页最新月 -> 与库 MAX 比较, 新则插入"""
        conn, cur = self.get_connection()
        cur.execute("SELECT MAX(source_month) AS m FROM market_production_monthly WHERE country_code='DE' AND data_source='vda_production'")
        row = cur.fetchone()
        max_m = row['m'] if isinstance(row, dict) else (row[0] if hasattr(row, '__getitem__') else None)
        # 抓统计页找最新月
        r = self._fetch(VDA_MONAT)
        if not r:
            return 0
        html = re.sub(r'<[^>]+>', ' ', r.text)
        # 找表头 'Juli 2026' 形式: 德语月份 + 年份
        m = re.search(r'(Januar|Februar|März|April|Mai|Juni|Juli|August|September|Oktober|November|Dezember)\s*(\d{4})', html)
        if not m:
            return 0
        latest = (int(m.group(2)), GER_MONTHS[m.group(1)])
        if max_m is None or date(*latest, 1) > max_m:
            # 从 Excel 全量重跑最稳 (含最新月)
            return self.crawl_archive()
        return 0


def main():
    import argparse
    ap = argparse.ArgumentParser()
    ap.add_argument('--incremental', action='store_true')
    ap.add_argument('--full', action='store_true')
    args = ap.parse_args()
    c = VdaCrawler()
    if args.incremental:
        print(c.crawl_incremental())
    elif args.full:
        print(c.crawl_archive())
    else:
        content = c.download_archive()
        data = c.parse_archive(content)
        print('months:', len(data), 'range:', min(data) if data else None, '~', max(data) if data else None)
        if data:
            print('latest 3:', sorted(data.items())[-3:])


if __name__ == '__main__':
    main()
