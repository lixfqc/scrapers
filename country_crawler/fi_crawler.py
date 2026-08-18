# -*- coding: utf-8 -*-
"""
aut.fi 芬兰汽车月度销量爬虫
数据源: https://aut.fi/markkinatilastot/kuukausitilastot/
数据类型: 品牌级月度首次注册(ENSIREKISTERÖINNIT) + BEV全品牌
口径: 首次注册 new registrations (Traficom源)
"""
import sys
sys.path.insert(0, '.')
import re
import io
import logging
import random
import requests
import psycopg2
from psycopg2.extras import RealDictCursor
from openpyxl import load_workbook
from datetime import datetime, date
from bs4 import BeautifulSoup
from base_crawler import BaseCrawler, DB_CONFIG

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s | %(levelname)s | %(message)s'
)

FI_BASE_URL = 'https://aut.fi'
FI_LIST_URL = f'{FI_BASE_URL}/markkinatilastot/kuukausitilastot/'

# 芬兰语月份名 -> 数字
FI_MONTH_MAP = {
    'tammikuu': 1, 'helmikuu': 2, 'maaliskuu': 3, 'huhtikuu': 4,
    'toukokuu': 5, 'kesakuu': 6, 'kesäkuu': 6, 'heinakuu': 7,
    'heinäkuu': 7, 'elokuu': 8, 'syyskuu': 9, 'lokakuu': 10,
    'marraskuu': 11, 'joulukuu': 12,
}


def _to_int(value):
    """安全整数转换"""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return int(value)
    s = str(value).strip().replace('\xa0', '').replace(',', '').replace(' ', '')
    try:
        return int(float(s))
    except (ValueError, TypeError):
        return None


class FinlandCrawler(BaseCrawler):
    """aut.fi 芬兰汽车月度销量爬虫"""

    def __init__(self):
        super().__init__(source_name='aut', country_code='FI')
        self.headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
        }
        self.session = requests.Session()
        self._file_index = None  # {month_key: url}

    # ============================================
    # 发现：索引页抓取全部月度xlsx链接
    # ============================================
    def discover_monthly_files(self):
        """从索引页抓取全部月度xlsx链接 -> {YYYY-MM: url}"""
        resp = self.session.get(FI_LIST_URL, headers=self.headers, timeout=30)
        if resp.status_code != 200:
            self.logger.error(f'索引页访问失败: HTTP {resp.status_code}')
            return {}
        soup = BeautifulSoup(resp.text, 'html.parser')
        files = {}
        for a in soup.find_all('a', href=True):
            href = a['href']
            if not re.search(r'\.xlsx?$', href, re.I):
                continue
            txt = a.get_text(strip=True)
            # 从链接文本解析月份名+年份，如 'heinäkuu_2026' / 'Joulukuu 2024' / 'Tammikuu_2023'
            m = re.search(r'([A-Za-zäöåÄÖÅ]+)[_\s]+(\d{4})', txt)
            if not m:
                # 兜底从URL提取
                m = re.search(r'/([A-Za-zäöåÄÖÅ]+?)[_0-9]*?(\d{4})\.xlsx', href, re.I)
            if not m:
                continue
            mon_name = m.group(1).lower()
            year = int(m.group(2))
            mon = FI_MONTH_MAP.get(mon_name)
            if not mon:
                continue
            if not href.startswith('http'):
                href = FI_BASE_URL + href
            key = f'{year}-{mon:02d}'
            files[key] = href
        self._file_index = files
        self.logger.info(f'发现 {len(files)} 个月度文件')
        return files

    def get_file_url(self, year, month):
        if self._file_index is None:
            self.discover_monthly_files()
        return self._file_index.get(f'{year}-{month:02d}')

    # ============================================
    # 下载
    # ============================================
    def download_excel(self, url):
        resp = self.session.get(url, headers=self.headers, timeout=60)
        if resp.status_code == 200 and len(resp.content) > 1000:
            self.logger.info(f'下载成功: {len(resp.content)} bytes')
            return resp.content
        self.logger.error(f'下载失败: HTTP {resp.status_code}, size={len(resp.content)}')
        return None

    # ============================================
    # 解析
    # ============================================
    def parse_excel(self, excel_content, year, month):
        """解析月度xlsx：Ha 30 merkkiä(品牌Top30) + BEV Ha merkit(纯电全品牌)"""
        records = []
        try:
            wb = load_workbook(io.BytesIO(excel_content), data_only=True)
            month_prefix = f'{month:02d}/{year}'

            # ---- Ha 30 merkkiä: 品牌Top30 ----
            if 'Ha 30 merkkiä' in wb.sheetnames:
                ws = wb['Ha 30 merkkiä']
                # 定位表头行（含 'Merkki' 且下一行有当月日期）
                header_idx = None
                for i in range(1, 15):
                    row = [ws.cell(row=i, column=c).value for c in range(1, 12)]
                    if row and any(isinstance(v, str) and v.strip() == 'Merkki' for v in row):
                        header_idx = i
                        break
                if header_idx:
                    # 当月销量列 = 表头行中含 YYYY 的列（07/2026）
                    qty_col = None
                    for c in range(1, 12):
                        v = ws.cell(row=header_idx, column=c).value
                        if isinstance(v, str) and re.search(r'\d{2}/\d{4}', v):
                            qty_col = c
                            break
                    if not qty_col:
                        qty_col = 3  # 兜底 col3
                    for i in range(header_idx + 2, ws.max_row + 1):
                        brand = ws.cell(row=i, column=2).value
                        if not brand or not str(brand).strip():
                            continue
                        brand_clean = re.sub(r'\s+', ' ', str(brand).replace('\xa0', ' ')).strip()
                        qty = _to_int(ws.cell(row=i, column=qty_col).value)
                        if qty is None:
                            continue
                        # 跳过汇总/合计/聚合行（Henkilöautotyhteensä 市场总量 / Muut merkit 其他品牌 / Matkailuautot 房车类别）
                        bu = brand_clean.upper()
                        if bu in ('YHTEENSÄ', 'YHTEENSA', 'TOTAL', 'MUUT MERKIT', 'MATKAILUAUTOT'):
                            continue
                        if re.search(r'HENKILÖ?AUTO.*YHTEEN|YHTEEN.*HENKILÖ?AUTO', bu):
                            continue
                        records.append(self._make_record(brand_clean, None, qty, year, month, 'Ha 30 merkkiä'))

            # ---- BEV Ha merkit: 纯电全品牌（车型级） ----
            bev_sheet = None
            for name in wb.sheetnames:
                if 'BEV' in name and 'merkit' in name.lower():
                    bev_sheet = name
                    break
            if bev_sheet:
                ws = wb[bev_sheet]
                # 找 'Kuukausi' 表头行，定位 M{month} 列
                month_col = None
                header_row_idx = None
                for i in range(1, 20):
                    row = [ws.cell(row=i, column=c).value for c in range(1, 25)]
                    if row and any(isinstance(v, str) and v.strip() == 'Kuukausi' for v in row):
                        header_row_idx = i
                        for c in range(1, 25):
                            v = ws.cell(row=i, column=c).value
                            if isinstance(v, str) and v.strip() == f'M{month}':
                                month_col = c
                                break
                        break
                if header_row_idx and month_col:
                    cur_brand = None
                    for i in range(header_row_idx + 2, ws.max_row + 1):
                        brand = ws.cell(row=i, column=1).value
                        model = ws.cell(row=i, column=2).value
                        qty = _to_int(ws.cell(row=i, column=month_col).value)
                        if qty is None:
                            continue
                        if brand and str(brand).strip():
                            cur_brand = re.sub(r'\s+', ' ', str(brand).replace('\xa0', ' ')).strip()
                        if not cur_brand:
                            continue
                        if model and str(model).strip():
                            records.append(self._make_record(cur_brand, str(model).strip(), qty, year, month, 'BEV Ha merkit'))
            wb.close()
        except Exception as e:
            self.logger.error(f'解析失败: {e}')
            import traceback
            traceback.print_exc()
        return records

    def _make_record(self, brand, model, qty, year, month, sheet_src):
        """构建入库记录"""
        return {
            'country_code': 'FI',
            'source_month': date(year, month, 1),
            'brand_name_raw': brand,
            'brand_id': None,
            'model_name': model,
            'vehicle_type': 'passenger',
            'energy_type': 'BEV' if sheet_src.startswith('BEV') else None,
            'segment': None,
            'raw_unit': 'units',
            'sales_volume_raw': qty,
            'sales_volume_normalized': qty,
            'revision_no': 1,
            'is_latest': True,
            'pub_date': None,
            'crawl_time': datetime.now(),
            'data_source': 'aut',
            'notes': f'FIN aut.fi ENSIREKISTERÖINNIT {sheet_src}',
        }

    # ============================================
    # 品牌匹配
    # ============================================
    def get_brand_id(self, brand_name_raw):
        brand_clean = str(brand_name_raw).strip().upper()
        try:
            conn = psycopg2.connect(**DB_CONFIG)
            cur = conn.cursor(cursor_factory=RealDictCursor)
            cur.execute("""
                SELECT id FROM brand_name_mapping
                WHERE UPPER(canonical_name) = %s OR UPPER(brand_name_cn) = %s
                ORDER BY id LIMIT 1
            """, (brand_clean, brand_clean))
            row = cur.fetchone()
            if row:
                cur.close(); conn.close()
                return row['id']
            cur.execute("""
                SELECT brand_id FROM brand_name_variant
                WHERE UPPER(variant_name) = %s
                ORDER BY brand_id LIMIT 1
            """, (brand_clean,))
            row = cur.fetchone()
            cur.close(); conn.close()
            return row['brand_id'] if row else None
        except Exception as e:
            self.logger.error(f'品牌ID查询失败: {e}')
        return None

    def save_sales(self, record):
        record['brand_id'] = self.get_brand_id(record['brand_name_raw'])
        if record['brand_id']:
            self.logger.debug(f'品牌匹配成功: {record["brand_name_raw"]} -> {record["brand_id"]}')
        return super().save_sales(record)

    # ============================================
    # 爬取
    # ============================================
    def crawl_month(self, year, month):
        url = self.get_file_url(year, month)
        if not url:
            self.logger.error(f'{year}-{month:02d} 未找到文件链接')
            return {'records': 0}
        content = self.download_excel(url)
        if not content:
            return {'records': 0}
        records = self.parse_excel(content, year, month)
        saved = 0
        for rec in records:
            if self.save_sales(rec):
                saved += 1
        self.logger.info(f'{year}-{month:02d} 解析{len(records)}条, 保存{saved}条')
        return {'records': saved}

    def crawl_range(self, start_year, start_month, end_year, end_month):
        self.discover_monthly_files()
        results = {}
        year, month = start_year, start_month
        while (year, month) <= (end_year, end_month):
            key = f'{year}-{month:02d}'
            results[key] = self.crawl_month(year, month)
            month += 1
            if month > 12:
                month = 1
                year += 1
        return results


def main():
    crawler = FinlandCrawler()
    results = crawler.crawl_range(2021, 1, 2026, 7)
    for k, v in results.items():
        print(f'{k}: {v}')


if __name__ == '__main__':
    main()
