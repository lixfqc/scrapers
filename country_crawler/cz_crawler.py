# -*- coding: utf-8 -*-
"""捷克 CZ 爬虫：SDA-CIA 月度品牌级新登记数据"""
import re
import sys
import random
import requests
import io as _io
from datetime import datetime, date
from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from base_crawler import BaseCrawler, DB_CONFIG, UA_LIST

CZ_BASE_URL = 'https://www.sda-cia.cz'
CZ_REPO_URL = f'{CZ_BASE_URL}/repository-volnedostupna?lang=EN&y={{year}}'
CZ_DL_URL = f'{CZ_BASE_URL}/repositoryfile?id={{fid}}&rnd={{rnd}}'

# sheet 名（EN版 / CZ版）
BRAND_SHEETS = ('PC in month', 'OA za měsíc')
FUEL_SHEETS = ('PC Fuel in month', 'OA Paliva za měsíc')
SKIP_BRANDS = ('TOTAL', 'CELKEM', 'OTHERS', 'JINÉ', 'NEZAŘAZENÉ', 'NEZAŘAZENO', 'ACROSS', 'CELKOVĚ', 'OTHER')

CZ_FUEL_MAP = {
    'PETROL + EL': 'PHEV', 'PETROL + EL.': 'PHEV', 'BENZIN + EL': 'PHEV', 'BENZÍN + EL': 'PHEV',
    'DIESEL + EL': 'PHEV', 'NAFTA + EL': 'PHEV',
    'PETROL + CNG': 'OTHER', 'PETROL + LPG': 'OTHER', 'DIESEL + CNG': 'OTHER', 'NAFTA + CNG': 'OTHER',
    'PETROL + PLUG': 'PHEV', 'PLUG': 'PHEV',
    'BENZÍN': 'GASOLINE', 'BENZIN': 'GASOLINE', 'PETROL': 'GASOLINE',
    'NAFTA': 'DIESEL', 'DIESEL': 'DIESEL',
    'ELEKTRO': 'BEV', 'ELEKTRICKÝ': 'BEV', 'ELECTRIC': 'BEV', 'EL': 'BEV',
    'CNG': 'CNG', 'LPG': 'LPG',
    'VODÍK': 'FCEV', 'HYDROGEN': 'FCEV',
    'HYBRID': 'HEV',
    'OTHERS': 'OTHER', 'UNCLASSIFIED': 'OTHER', 'JINÉ': 'OTHER', 'OSTATNÍ': 'OTHER',
}

class CzechCrawler(BaseCrawler):
    def __init__(self):
        super().__init__('SDA_CIA', 'CZ')
        self.session = requests.Session()
        self.session.headers.update({
            'User-Agent': random.choice(UA_LIST),
            'Referer': CZ_BASE_URL,
        })
        self._brand_id_cache = {}

    def _safe_int(self, v):
        try:
            if v is None:
                return 0
            if isinstance(v, float):
                return int(v)
            s = str(v).replace('\xa0', '').replace(' ', '').replace(',', '').strip()
            if not s or s in ('-', 'nan', 'None'):
                return 0
            return int(float(s))
        except Exception:
            return 0

    def _get_repo_page(self, year):
        r = self.retry_request(self.session.get, CZ_REPO_URL.format(year=year), timeout=30)
        r.raise_for_status()
        return r.text

    def _get_download_url(self, year, month):
        """从仓库列表页提取 {year}-{month}.monthly.EN.xlsx 的下载链接"""
        html = self._get_repo_page(year)
        pat = re.compile(
            r'href="([^"]*repositoryfile\?id=(\d+)[^"]*)"[^>]*>\s*'
            + str(year) + r'-' + str(month) + r'\.monthly\.EN\.xlsx',
            re.I,
        )
        for m in pat.finditer(html):
            href = m.group(1).replace('&amp;', '&')
            if href.startswith('http'):
                return href
            return CZ_BASE_URL + '/' + href.lstrip('/')
        return None

    def discover_download_url(self, year, month):
        return self._get_download_url(year, month)

    def download_excel(self, url):
        r = self.retry_request(self.session.get, url, timeout=60)
        r.raise_for_status()
        return r.content

    def _load_brand_sheet(self, content):
        """读入品牌级sheet，返回 (品牌行列表, 表头月份, total)"""
        try:
            import io as _i
            wb = load_workbook(_i.BytesIO(content), data_only=True)
        except Exception as e:
            raise ValueError(f'无法解析xlsx: {e}')

        ws = None
        for name in BRAND_SHEETS:
            for sn in wb.sheetnames:
                if sn.strip() == name or sn.strip().startswith(name):
                    ws = wb[sn]
                    break
            if ws:
                break
        if ws is None:
            raise ValueError('未找到品牌级sheet')

        # 表头 r2 含 Period (7/2026)，r3=Make|Units|Share|Position
        period_ym = None
        for r in range(1, 5):
            for c in range(1, 5):
                v = ws.cell(r, c).value
                if isinstance(v, str) and re.match(r'\d{1,2}\s*/\s*\d{4}', v.replace('\xa0', ' ')):
                    m = re.match(r'(\d{1,2})\s*/\s*(\d{4})', v.replace('\xa0', ' '))
                    if m:
                        period_ym = (int(m.group(2)), int(m.group(1)))
                    break
            if period_ym:
                break

        # 定位表头行（Make/Značka 与 Units/Ks 列）
        header_idx = None
        brand_col = 0
        qty_col = 2
        for r in range(1, 8):
            for c in range(1, ws.max_column + 1):
                v = ws.cell(r, c).value
                if isinstance(v, str) and v.strip().lower() in ('make', 'značka', 'znacka', 'marca'):
                    header_idx = r
                    brand_col = c - 1
                    # 找 Units/Ks 列
                    for cc in range(c, min(c + 6, ws.max_column) + 1):
                        vv = ws.cell(r, cc).value
                        if isinstance(vv, str) and vv.strip().lower() in ('units', 'ks', 'počet', 'pocet', 'count'):
                            qty_col = cc - 1
                            break
                    if brand_col == 0:
                        brand_col = 0
                    break
            if header_idx is not None:
                break
        if header_idx is None:
            raise ValueError('未定位表头行')

        brands = []
        total = None
        for r in range(header_idx + 1, ws.max_row + 1):
            brand = ws.cell(r, brand_col + 1).value
            qty = self._safe_int(ws.cell(r, qty_col + 1).value)
            if brand is None or not str(brand).strip():
                continue
            b = str(brand).strip()
            if b.upper() in ('TOTAL', 'CELKEM') or b.upper().startswith('TOTAL'):
                if qty > 0:
                    total = qty
                    break
                continue
            if b.upper() in SKIP_BRANDS or b.upper().startswith('OTHER') or b.upper().startswith('JINÉ'):
                continue
            brands.append((b, qty))
        return brands, period_ym, total

    def _load_fuel_sheet(self, content):
        """读品牌×燃料sheet，返回 [(brand, fuel, qty)]"""
        import io as _i
        wb = load_workbook(_i.BytesIO(content), data_only=True)
        ws = None
        for name in FUEL_SHEETS:
            for sn in wb.sheetnames:
                if sn.strip() == name or sn.strip().startswith(name):
                    ws = wb[sn]
                    break
            if ws:
                break
        if ws is None:
            return []
        # r3=燃料名（0-based偶数列），r5起数据：col1=品牌
        HEADER_LABELS = {'FUEL', 'PALIVO', 'MAKE', 'BRAND', 'ZNAČKA', 'ZNACKA', 'TOTAL', 'SHARE', 'PODÍL', 'KS', 'UNITS'}
        fuel_cols = []
        header_idx = None
        for r in range(1, 6):
            for c in range(2, min(28, ws.max_column + 1)):
                v = ws.cell(r, c).value
                if isinstance(v, str) and v.strip() and (c - 1) % 2 == 0:
                    fname = v.strip().upper()
                    if fname in HEADER_LABELS:
                        continue
                    mapped = None
                    # 优先匹配最长的复合燃料名（如 'PETROL + EL.' > 'PETROL'）
                    for k in sorted(CZ_FUEL_MAP, key=len, reverse=True):
                        if k in fname:
                            mapped = CZ_FUEL_MAP[k]
                            break
                    if mapped:
                        fuel_cols.append((c - 1, mapped))
            if len(fuel_cols) >= 3:
                header_idx = r
                break
        records = []
        for r in range(header_idx + 1, ws.max_row + 1):
            brand = ws.cell(r, 1).value
            if brand is None or not str(brand).strip():
                continue
            b = str(brand).strip()
            if b.upper() in SKIP_BRANDS or b.upper() in ('TOTAL', 'CELKEM', 'SHARE') or b.upper().startswith('TOTAL'):
                continue
            for col, fuel in fuel_cols:
                qty = self._safe_int(ws.cell(r, col + 1).value)
                if qty > 0:
                    records.append((b, fuel, qty))
        return records

    def get_brand_id(self, brand_raw):
        if brand_raw in self._brand_id_cache:
            return self._brand_id_cache[brand_raw]
        conn, cur = self.get_connection()
        lookup = brand_raw.strip().upper()
        cur.execute("SELECT id FROM brand_name_mapping WHERE UPPER(canonical_name)=%s OR UPPER(brand_name_cn)=%s", (lookup, lookup))
        row = cur.fetchone()
        if not row:
            cur.execute("SELECT brand_id FROM brand_name_variant WHERE UPPER(variant_name)=%s", (lookup,))
            row = cur.fetchone()
        bid = None
        if row:
            if 'id' in row and row['id'] is not None:
                bid = row['id']
            elif 'brand_id' in row and row['brand_id'] is not None:
                bid = row['brand_id']
        self._brand_id_cache[brand_raw] = bid
        return bid

    def save_sales(self, record):
        record['brand_id'] = self.get_brand_id(record['brand_name_raw'])
        return super().save_sales(record)

    def crawl_month(self, year, month):
        url = self.discover_download_url(year, month)
        if not url:
            return {'records': 0, 'msg': f'no URL for {year}-{month}'}
        content = self.download_excel(url)
        brands, period_ym, total = self._load_brand_sheet(content)
        if period_ym and period_ym != (year, month):
            return {'records': 0, 'msg': f'file period {period_ym} != target ({year},{month})'}
        saved = 0
        for brand, qty in brands:
            record = {
                'country_code': 'CZ', 'source_month': date(year, month, 1),
                'brand_name_raw': brand, 'model_name': None,
                'vehicle_type': 'passenger', 'energy_type': None, 'segment': None,
                'raw_unit': 'units', 'sales_volume_raw': qty,
                'sales_volume_normalized': qty, 'revision_no': 1, 'is_latest': True,
                'pub_date': None, 'crawl_time': datetime.now(), 'data_source': 'SDA_CIA',
                'notes': 'SDA-CIA new PC registrations by make (monthly)',
            }
            self.save_sales(record)
            saved += 1
        # 燃料级（品牌×能源）
        try:
            fuel_rows = self._load_fuel_sheet(content)
            for brand, fuel, qty in fuel_rows:
                record = {
                    'country_code': 'CZ', 'source_month': date(year, month, 1),
                    'brand_name_raw': brand, 'model_name': None,
                    'vehicle_type': 'passenger', 'energy_type': fuel, 'segment': None,
                    'raw_unit': 'units', 'sales_volume_raw': qty,
                    'sales_volume_normalized': qty, 'revision_no': 1, 'is_latest': True,
                    'pub_date': None, 'crawl_time': datetime.now(), 'data_source': 'SDA_CIA',
                    'notes': 'SDA-CIA new PC registrations by make x fuel (monthly)',
                }
                self.save_sales(record)
                saved += 1
        except Exception as e:
            print(f'燃料级解析失败: {e}')
        return {'records': saved, 'total': total}

    def _get_db_max_month(self):
        conn, cur = self.get_connection()
        cur.execute("SELECT MAX(source_month) AS m FROM market_sales_monthly WHERE country_code='CZ'")
        row = cur.fetchone()
        return row['m'] if row and row['m'] else None

    def latest_available_month(self):
        """探测最新可用的年月（从当年列表页找最大月）"""
        year = datetime.now().year
        for y in (year, year - 1):
            try:
                html = self._get_repo_page(y)
                months = [int(m) for m in re.findall(str(y) + r'-(\d{1,2})\.monthly\.EN\.xlsx', html, re.I)]
                if months:
                    return (y, max(months))
            except Exception:
                continue
        return None

    def crawl_incremental(self):
        latest = self.latest_available_month()
        if not latest:
            return 0
        max_m = self._get_db_max_month()
        latest_date = date(latest[0], latest[1], 1)
        if max_m and latest_date <= max_m:
            return 0
        # 从库MAX+1到最新逐月
        y, m = (max_m.year, max_m.month + 1) if max_m else (latest[0], 1)
        if m > 12:
            y += 1; m = 1
        total = 0
        while (y, m) <= latest:
            try:
                res = self.crawl_month(y, m)
                total += res['records']
            except Exception as e:
                print(f'CZ {y}-{m} 失败: {e}')
            m += 1
            if m > 12:
                y += 1; m = 1
        return total

    def crawl_range(self, start_year, start_month, end_year, end_month):
        total = 0
        y, m = start_year, start_month
        while (y, m) <= (end_year, end_month):
            try:
                res = self.crawl_month(y, m)
                total += res['records']
                print(f'CZ {y}-{m}: {res["records"]}条')
            except Exception as e:
                print(f'CZ {y}-{m} 失败: {e}')
            m += 1
            if m > 12:
                y += 1; m = 1
        return total


if __name__ == '__main__':
    c = CzechCrawler()
    c.crawl_range(2022, 1, 2026, 7)
