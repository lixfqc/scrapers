# -*- coding: utf-8 -*-
"""
SE 瑞典汽车月度销量爬虫
数据源: https://mobilitysweden.se/ (Bil Sweden / Mobility Sweden)
数据类型: 品牌级+车型级乘用车新注册 (Nyregistreringar, Personbilar)
数据来源: Transportstyrelsen (瑞典运输署)

站点结构:
- 列表页: https://mobilitysweden.se/statistik/Nyregistreringar_per_manad_1 (?page=1/2, 2015-2026)
- 新闻稿URL: .../nyregistreringar-{YYYY}/{slug}
- 详情页附件: /storage/{HASH}/xlsx/media/{HASH}/Månadsrapport Nyregistreringar {månad} {YYYY}.xlsx

Excel结构 ('PB - Fabrikat och modell' / 'A.2 Fabrikat och modeller PB' / 'Registrations'):
- row8 = 表头 (D=Fabrikat, F=Modell, J=月份名)
- row10+ 数据: D列=品牌(汇总行有值), F列=车型(汇总行='Total', 末行='Totalt')
- J列(索引9)=当月注册量, N列(索引13)=YTD, O列(索引14)=YTD LY
- 品牌块 = 品牌名行(F='Total') + 该品牌车型行
"""
import sys
sys.path.insert(0, '.')
import os
import re
import io
import json
import logging
import requests
import openpyxl
from datetime import datetime, date
from bs4 import BeautifulSoup
from base_crawler import BaseCrawler, DB_CONFIG

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s | %(levelname)s | %(message)s'
)

SE_BASE_URL = 'https://mobilitysweden.se'
SE_LIST_URL = f'{SE_BASE_URL}/statistik/Nyregistreringar_per_manad_1'

# 瑞典语月份名 -> 月份
SWE_MONTH_MAP = {
    'januari': 1, 'februari': 2, 'mars': 3, 'april': 4, 'maj': 5,
    'juni': 6, 'juli': 7, 'augusti': 8, 'september': 9, 'oktober': 10,
    'november': 11, 'december': 12,
}
SWE_MONTH_REV = {v: k for k, v in SWE_MONTH_MAP.items()}

# 主sheet名（按版本）
SHEET_NAMES = [
    'PB - Fabrikat och modell',
    'A.2 Fabrikat och modeller PB',
    'Registrations',
]

# 品牌聚合/总量行（F列值判断）
SKIP_MODELS = ('TOTALT', 'TOTAL', 'SUMMA', 'ÖVRIGA', 'OVRIGA')


class SwedenCrawler(BaseCrawler):
    """Mobility Sweden 瑞典汽车销量爬虫"""

    def __init__(self):
        super().__init__(source_name='mobilitysweden', country_code='SE')
        self.headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Accept-Language': 'sv-SE,sv;q=0.9,en;q=0.8',
        }
        self.session = requests.Session()

    # ---------- 列表抓取 ----------
    def crawl_list(self):
        """抓取全部新闻稿列表, 返回 [(url, title, date_str)]"""
        items = []
        for page in (1, 2):
            url = f'{SE_LIST_URL}?page={page}'
            resp = self.session.get(url, headers=self.headers, timeout=30)
            if resp.status_code != 200:
                self.logger.error(f'列表页失败: {url} HTTP {resp.status_code}')
                continue
            resp.encoding = resp.apparent_encoding
            soup = BeautifulSoup(resp.text, 'html.parser')
            for item in soup.find_all('li', class_='news-list__item'):
                h2 = item.find('h2', class_='list_header')
                if not h2:
                    continue
                a = h2.find('a')
                if not a or not a.get('href'):
                    continue
                p = item.find('p', class_='list_date')
                date_str = p.get_text(strip=True) if p else ''
                items.append((a['href'], a.get_text(strip=True), date_str))
            self.logger.info(f'列表页 {page} 抓到 {len(items)} 项')
        return items

    # ---------- 新闻稿详情 → 附件URL ----------
    def find_excel_attachment(self, detail_url):
        """从新闻稿详情页提取主报告 xlsx 附件URL (品牌+车型表)"""
        resp = self.session.get(detail_url, headers=self.headers, timeout=30)
        if resp.status_code != 200:
            self.logger.error(f'详情页失败: {detail_url} HTTP {resp.status_code}')
            return None
        resp.encoding = resp.apparent_encoding
        for m in re.finditer(r'href="([^"]+\.(?:xlsx|xls))"', resp.text, re.I):
            href = m.group(1)
            low = href.lower()
            if not low.endswith('.xlsx'):
                continue
            # 主报告: 文件名含 Nyregistreringar 或 Månadsrapport
            has_main = ('nyregistreringar' in low or 'nadsrapport' in low)
            if not has_main:
                continue
            # 排除地区/卡车/客车/商用车/车队文件
            excluded = ('län' in low or 'l%c3%a4n' in low or 'kommun' in low
                        or 'topplista' in low or 'lastbilar' in low
                        or 'bussar' in low or 'l%c3%a4tta' in low
                        or 'transport' in low or 'personbilar per' in low
                        or 'personbilar%20per' in low or 'per%20l%c3%a4n' in low
                        or 'tunga' in low or 'drivmedelsfo' in low
                        or 'bransch' in low or 'elbil' in low)
            if excluded:
                continue
            if href.startswith('http'):
                return href
            return SE_BASE_URL + href
        self.logger.warning(f'未找到xlsx附件: {detail_url}')
        return None

    # ---------- 解析 ----------
    def parse_excel(self, excel_content, source_month):
        """解析主sheet, 返回 [(brand, model, qty)]"""
        rows_out = []
        try:
            wb = openpyxl.load_workbook(io.BytesIO(excel_content), data_only=True, read_only=True)
        except Exception as e:
            self.logger.error(f'xlsx打开失败: {e}')
            return rows_out

        ws = None
        for name in SHEET_NAMES:
            if name in wb.sheetnames:
                ws = wb[name]
                self.logger.info(f'解析Sheet: {name}')
                break
        if ws is None:
            self.logger.warning(f'无主sheet, 可用: {wb.sheetnames[:10]}')
            wb.close()
            return rows_out

        data = list(ws.iter_rows(values_only=True))
        wb.close()

        # 探测表头行与列位置
        # 2026版: 品牌=D(3), 车型=F(5), 数量=J(9), 品牌行F='Total', 表头行D='Fabrikat (Antal reg. YTD↓)'
        # 2024版: 品牌+车型同列B(1)('Fabrikat / modell'), 数量=E(4), 新品牌行A列='1(1)'
        # 注意: 前几行可能有标题 'Fabrikat och modell -' 也含FABRIKAT, 需排除(标题行无'MODELL'单独列或下方无数据)
        header_idx = None
        fab_col = None
        merged_col = False   # 品牌与车型同一列
        qty_col = None
        for r_idx, row in enumerate(data[:20]):
            if row is None:
                continue
            row_txt = {c_idx: str(cell).upper() for c_idx, cell in enumerate(row) if cell is not None}
            fab_positions = [c for c, t in row_txt.items() if 'FABRIKAT' in t]
            if not fab_positions:
                continue
            # 排除标题行 'Fabrikat och modell -': 真实表头单元格含 '(' 或 '/' (如 'Fabrikat (Antal reg. YTD' / 'Fabrikat / modell')
            real_header = any('(' in row_txt[c] or '/' in row_txt[c] for c in fab_positions)
            if not real_header:
                continue
            # 2024版: 同单元格含 'MODELL' (如 'Fabrikat / modell')
            merged_cells = [c for c in fab_positions if 'MODELL' in row_txt[c]]
            # 2026版: 表头行另有独立 'MODELL' 列
            has_model_col = any('MODELL' in t for c, t in row_txt.items())
            if not merged_cells and not has_model_col:
                continue   # 标题行(仅 'Fabrikat och modell -') 无独立MODELL列, 跳过
            header_idx = r_idx
            fab_col = fab_positions[0]
            if merged_cells:
                merged_col = True
                fab_col = merged_cells[0]
            # 数量列: 表头行找纯年份 20xx (2024版) 或月份名列 (2026版)
            for c2 in range(fab_col + 1, len(row)):
                v = row[c2]
                if v is None:
                    continue
                vs = str(v).strip()
                if re.fullmatch(r'20\d{2}', vs):
                    qty_col = c2
                    break
            if qty_col is None:
                # 2026版: 表头最后一个月份名(juli/juni等)列即当月数量列
                for c2 in range(len(row) - 1, fab_col, -1):
                    v = row[c2]
                    if v is None:
                        continue
                    vs = str(v).strip().lower()
                    if vs in SWE_MONTH_MAP or vs.rstrip('.') in SWE_MONTH_MAP:
                        qty_col = c2
                        break
            break

        if header_idx is None:
            self.logger.warning(f'未找到FABRIKAT表头, 可用前20行')
            return rows_out

        header_row = data[header_idx]
        if not merged_col:
            # 2026版: 车型列 = 表头含MODELL的列(品牌列之后)
            model_col = None
            for c2 in range(fab_col + 1, min(fab_col + 5, len(header_row))):
                if header_row[c2] is not None and 'MODELL' in str(header_row[c2]).upper():
                    model_col = c2
                    break
            if model_col is None:
                model_col = fab_col + 2  # 默认 F
            if qty_col is None:
                qty_col = fab_col + 6    # 默认 J
        else:
            # 2024版: 品牌车型同列; 数量列若非年份则默认E(4)
            model_col = None
            if qty_col is None:
                qty_col = 4

        self.logger.info(f'表头行={header_idx} 品牌列={fab_col} 车型列={model_col} 数量列={qty_col} merged={merged_col}')

        current_brand = None
        start_idx = header_idx + 1
        for r_idx, row in enumerate(data):
            if r_idx < start_idx:
                continue
            if not row or len(row) <= qty_col:
                continue
            qty_raw = row[qty_col]
            qty = None
            if qty_raw is not None:
                try:
                    qty = int(qty_raw)
                except (ValueError, TypeError):
                    qty = None

            if merged_col:
                # 2024版: A列'数字(数字)'=新品牌行, B=品牌; 其余行B=车型
                rank = str(row[0]).strip() if row[0] is not None else ''
                bval = str(row[fab_col]).strip() if fab_col < len(row) and row[fab_col] is not None else ''
                if not bval:
                    continue
                if re.fullmatch(r'\d+\(\d+\)', rank):
                    if bval.upper() in ('TOTAL', 'TOTALT') or bval.upper() == 'ÖVRIGA':
                        current_brand = None
                        continue
                    current_brand = bval
                    continue
                # 车型行
                if current_brand and bval and bval.upper() not in SKIP_MODELS and not any(s in bval.upper() for s in ('ÖVRIGA', 'OVRIGA')):
                    if qty is not None:
                        rows_out.append((current_brand, bval, qty))
            else:
                # 2026版
                brand = str(row[fab_col]).strip() if fab_col < len(row) and row[fab_col] is not None else ''
                model = str(row[model_col]).strip() if model_col is not None and model_col < len(row) and row[model_col] is not None else ''
                if brand:
                    if model.upper() in ('TOTAL', 'TOTALT', 'SUMMA') or model == '':
                        if model.upper() == 'TOTALT' or brand.upper() == 'ÖVRIGA':
                            current_brand = None
                            continue
                        current_brand = brand
                    else:
                        current_brand = brand
                        if qty is not None and model:
                            rows_out.append((brand, model, qty))
                elif model and current_brand:
                    if model.upper() in SKIP_MODELS or any(s in model.upper() for s in ('ÖVRIGA', 'OVRIGA')):
                        continue
                    if qty is not None:
                        rows_out.append((current_brand, model, qty))

        self.logger.info(f'解析到 {len(rows_out)} 条车型级记录')
        return rows_out

    # ---------- 品牌ID ----------
    def get_brand_id(self, brand_name):
        """UPPER匹配 canonical_name/brand_name_cn, variant回退"""
        lookup = str(brand_name).strip().upper()
        try:
            conn, cur = self.get_connection()
            cur.execute("""
                SELECT id FROM brand_name_mapping
                WHERE UPPER(canonical_name) = %s OR UPPER(brand_name_cn) = %s
                LIMIT 1
            """, (lookup, lookup))
            row = cur.fetchone()
            if row:
                return row['id']
            cur.execute("""
                SELECT brand_id FROM brand_name_variant
                WHERE UPPER(variant_name) = %s
                LIMIT 1
            """, (lookup,))
            row = cur.fetchone()
            if row:
                return row['brand_id']
        except Exception as e:
            self.logger.error(f'品牌ID查询失败: {e}')
        return None

    def save_sales(self, record):
        record['brand_id'] = self.get_brand_id(record['brand_name_raw'])
        return super().save_sales(record)

    # ---------- 主流程 ----------
    def crawl_month(self, year, month):
        """爬取指定月份 (通过列表日期/标题定位当月新闻稿)"""
        month_name = SWE_MONTH_REV[month]

        items = self.crawl_list()
        detail_url = None

        # 匹配候选: 含 nyregistreringar-{year} 前缀, 且标题含月份名
        candidates = []
        for url, title, date_str in items:
            low = url.lower()
            if f'nyregistreringar-{year}' in low:
                # 标题含月份关键词优先
                if month_name in title.lower():
                    candidates.append((url, title, date_str))
                # 日期兜底: 发布时间为下月1-5日 → 数据月份=发布月-1
                m = re.search(r'(\d{1,2})\s+([a-zåäö]+)\s+(\d{4})', date_str.lower())
                if m:
                    day, mon_name, mon_year = int(m.group(1)), m.group(2), int(m.group(3))
                    pub_month = SWE_MONTH_MAP.get(mon_name)
                    if pub_month:
                        data_month = pub_month - 1
                        data_year = mon_year
                        if data_month == 0:
                            data_month, data_year = 12, mon_year - 1
                        if (data_year, data_month) == (year, month):
                            candidates.append((url, title, date_str))

        # 去重并优先标题匹配的
        seen = set()
        picked = []
        for c in candidates:
            if c[0] not in seen:
                seen.add(c[0])
                picked.append(c)

        for url, title, date_str in picked:
            # 检查详情页是否含 Månadsrapport xlsx
            excel_url = self.find_excel_attachment(url)
            if excel_url:
                detail_url = url
                self.logger.info(f'命中新闻稿: {url}')
                break

        if not detail_url or not excel_url:
            self.logger.error(f'{year}-{month:02d} 未找到含附件新闻稿 (候选{len(picked)})')
            return {'records': 0}

        resp = self.session.get(excel_url, headers=self.headers, timeout=60)
        if resp.status_code != 200 or len(resp.content) < 1000:
            self.logger.error(f'下载失败: HTTP {resp.status_code}, size={len(resp.content)}')
            return {'records': 0}

        records = self.parse_excel(resp.content, date(year, month, 1))
        saved = 0
        for brand, model, qty in records:
            record = {
                'country_code': 'SE',
                'source_month': date(year, month, 1),
                'brand_name_raw': brand,
                'brand_id': None,
                'model_name': model,
                'vehicle_type': 'passenger',
                'energy_type': None,
                'segment': None,
                'raw_unit': 'units',
                'sales_volume_raw': qty,
                'sales_volume_normalized': qty,
                'revision_no': 1,
                'is_latest': True,
                'pub_date': None,
                'crawl_time': datetime.now(),
                'data_source': 'mobilitysweden',
                'notes': f'Mobility Sweden Nyregistreringar {year}-{month:02d}',
            }
            if self.save_sales(record):
                saved += 1

        self.logger.info(f'{year}-{month:02d} 保存 {saved} 条')
        return {'records': saved}

    def crawl_range(self, start_year, start_month, end_year, end_month):
        """爬取范围, 逐月"""
        results = {}
        cy, cm = start_year, start_month
        while (cy, cm) <= (end_year, end_month):
            key = f'{cy}-{cm:02d}'
            try:
                results[key] = self.crawl_month(cy, cm)
            except Exception as e:
                self.logger.error(f'{key} 异常: {e}')
                import traceback
                traceback.print_exc()
                results[key] = {'records': 0, 'error': str(e)}
            cm += 1
            if cm > 12:
                cm = 1
                cy += 1
        return results


def main():
    """主函数: 爬取瑞典数据"""
    crawler = SwedenCrawler()
    print('=== Mobility Sweden 瑞典汽车销量爬虫 ===')
    results = crawler.crawl_range(2024, 1, 2026, 7)
    for k, v in results.items():
        print(f'  {k}: {v}')
    print('完成!')


if __name__ == '__main__':
    main()
