# -*- coding: utf-8 -*-
"""
KAMA 韩国汽车销量爬虫（2026-08-13 重写版）
数据源: https://www.kama.or.kr/  자동차등록통계 (Vehicle Registration Statistics)

数据口径: 신규등록 (当月新登记 = 当月注册增量)
- Sheet NR{YYYY}{MM}_2: 제작사별 차종별 사용연료별 신규등록현황
  - r4: 月份标题 (YYYY년 M월)
  - r5: 表头 구분|차종|합계|휘발유|경유|ＬＰＧ|CNG|하이브리드|전기|수소|기타
  - r6: 총계 行业合计行; r7-10: 行业车型行; r11起: 品牌块(品牌|합계 合计行 + 车型行)

存储口径 (受 market_sales_monthly 唯一约束 UNIQUE(country_code, source_month,
brand_name_raw, model_name, energy_type, revision_no) 限制，不含 vehicle_type):
1. 品牌级: 品牌합계行 col2 → brand=韩语品牌, energy=None, vehicle=None
2. 品牌×燃料级: 品牌합계行各燃料列 → brand=韩语品牌, energy=燃料
3. 行业燃料级: 行业총계行各燃料列 → brand='KOREA INDUSTRY', energy=燃料

文件发现: 遍历 cmd=L 列表页 (pagenum=1..51, 每页10条), 文件名 R{YYYY}{MM}.xlsx
即当月数据 (容忍 -1/_수정판/_V2 后缀; 文件内 sheet 名可能残留模板月份, 以文件名月份为准)

下载机制:
- POST https://www.kama.or.kr/jsp/common/FileDown.jsp
- 参数: org_fileName, server_fileName, boardmaster_id, path
"""
import sys
sys.path.insert(0, '.')
import io
import re
import time
import random
import logging
import requests
import psycopg2
from psycopg2.extras import RealDictCursor
import openpyxl
from datetime import datetime, date
from base_crawler import BaseCrawler, DB_CONFIG

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s | %(levelname)s | %(message)s'
)

# ============================================
# 韩语品牌映射表
# ============================================
KOREAN_BRAND_MAP = {
    # 韩国本土品牌
    '현대': 'HYUNDAI',
    '기아': 'KIA',
    '한국지엠': 'GM KOREA',
    '르노코리아': 'RENAULT KOREA',
    '르노 KOR': 'RENAULT KOREA',
    'KG모빌리티': 'KG MOBILITY',
    '대우버스': 'DAEWOO BUS',
    '타타대우': 'TATA DAEWOO',
    '쌍용': 'SSANGYONG',
    '쌍용자동차': 'SSANGYONG',
    '르노삼성': 'RENAULT SAMSUNG',
    '지엠대우': 'GM DAEWOO',
    '대우자동차': 'DAEWOO',
    'CT&T': 'CT&T',
    'KG': 'KG',
    '광림': 'KWANGRIM',
    '기타': 'OTHER',
    '대우': 'DAEWOO',
    '대우중': 'DAEWOO HEAVY',
    '삼성상': 'SAMSUNG COMMERCIAL',
    '수산': 'SUSAN',
    '수입': 'IMPORTED',
    '아시아': 'ASIA',
    'AD 모터스': 'AD MOTORS',
    '진도': 'JINDO',
    '합계': 'TOTAL',
    '총계': 'TOTAL',
    '계': 'TOTAL',
    '소계': 'SUBTOTAL',
}

# 燃料类型映射
FUEL_TYPE_MAP = {
    '휘발유': 'GASOLINE',
    '경유': 'DIESEL',
    '전기': 'BEV',
    '하이브리드': 'HEV',
    '플러그인하이브리드': 'PHEV',
    '수소': 'FCEV',
    'LPG': 'LPG',
    'LP가스': 'LPG',
    'ＬＰＧ': 'LPG',
    'CNG': 'CNG',
    '가스': 'GAS',
    '기타': 'OTHER',
    '합계': 'TOTAL',
    '총계': 'TOTAL',
}

# KAMA网站配置
KAMA_BASE_URL = 'https://www.kama.or.kr'
KAMA_LIST_URL = f'{KAMA_BASE_URL}/NewsController?cmd=L&board_id=504&boardmaster_id=Register&menunum=0003'
KAMA_FILE_DOWNLOAD_URL = f'{KAMA_BASE_URL}/jsp/common/FileDown.jsp'
KAMA_MAX_PAGES = 60

# 车型映射
VEHICLE_TYPE_MAP = {
    '승용차': 'passenger',
    '승합차': 'van',
    '화물차': 'commercial',
    '특수차': 'special',
}


def _map_korean_brand(brand_raw):
    """映射韩语品牌名到英文"""
    brand_clean = str(brand_raw).strip()

    if brand_clean in KOREAN_BRAND_MAP:
        return KOREAN_BRAND_MAP[brand_clean]

    for kr_name, en_name in KOREAN_BRAND_MAP.items():
        if kr_name in brand_clean:
            return en_name

    return brand_clean


def _map_fuel_type(fuel_raw):
    """映射韩语燃料类型到标准类型"""
    fuel_clean = str(fuel_raw).strip()

    if fuel_clean in FUEL_TYPE_MAP:
        return FUEL_TYPE_MAP[fuel_clean]

    for kr_name, en_name in FUEL_TYPE_MAP.items():
        if kr_name in fuel_clean:
            return en_name

    return fuel_clean


def _to_int(value):
    """安全转换整数"""
    if value is None:
        return 0
    try:
        if isinstance(value, float):
            return int(value)
        s = str(value).replace(',', '').replace(' ', '').strip()
        if not s:
            return 0
        return int(s)
    except (ValueError, TypeError):
        return 0


class KamaCrawler(BaseCrawler):
    """KAMA韩国汽车销量爬虫"""

    def __init__(self):
        super().__init__(source_name='kama', country_code='KR')
        self.headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Referer': KAMA_BASE_URL,
        }
        self.session = requests.Session()
        self._file_index = None

    # ------------------------------------------------------------
    # 文件发现：遍历 cmd=L 列表页构建全量文件索引
    # ------------------------------------------------------------
    def _build_file_index(self):
        """遍历列表页，构建 {org_fileName: params} 索引"""
        if self._file_index:
            return self._file_index

        index = {}
        pagenum = 1
        total_pages = 1
        self.logger.info('构建KAMA文件索引...')

        while pagenum <= KAMA_MAX_PAGES:
            try:
                if pagenum == 1:
                    resp = self.session.get(KAMA_LIST_URL, headers=self.headers, timeout=30)
                else:
                    resp = self.session.post(
                        KAMA_LIST_URL,
                        data={'cmd': 'L', 'board_id': '504', 'boardmaster_id': 'Register',
                              'menunum': '0003', 'pagenum': str(pagenum)},
                        headers=self.headers,
                        timeout=30
                    )
                if resp.status_code != 200:
                    self.logger.error(f'列表页 {pagenum} HTTP {resp.status_code}')
                    break
                resp.encoding = 'euc-kr'
                html = resp.text

                # 从第一页提取 movePage 的最大页码作为总页数
                if pagenum == 1:
                    for m in re.finditer(r"movePage\(\s*'?(\d+)'?\s*\)", html):
                        total_pages = max(total_pages, int(m.group(1)))
                    self.logger.info(f'列表共 {total_pages} 页')

                added = 0
                for m in re.finditer(r"fileDown\('([^']+)','([^']+)','([^']+)','([^']+)'\)", html):
                    org, server, path, bm = m.groups()
                    if org not in index:
                        index[org] = {
                            'org_fileName': org,
                            'server_fileName': server,
                            'path': path,
                            'boardmaster_id': bm,
                        }
                        added += 1

                self.logger.info(f'列表页 {pagenum}: 新增 {added} 个文件 (累计 {len(index)})')
                if added == 0:
                    break
                pagenum += 1
                if pagenum > total_pages:
                    break
                time.sleep(random.uniform(1, 2))
            except Exception as e:
                self.logger.error(f'列表页 {pagenum} 请求异常: {e}')
                break

        self._file_index = index
        self.logger.info(f'文件索引构建完成: {len(index)} 个文件')
        return index

    def latest_available_month(self):
        """快速探测KAMA最新可用月份（只抓列表第1页最新10条，不构建全量索引）

        Returns:
            (int, int) or None: (year, month) 最新可用月份
        """
        try:
            resp = self.session.get(KAMA_LIST_URL, headers=self.headers, timeout=30)
            if resp.status_code != 200:
                return None
            resp.encoding = 'euc-kr'
            latest = None
            for m in re.finditer(r"fileDown\('([^']+)','([^']+)','([^']+)','([^']+)'\)", resp.text):
                org = m.group(1)
                mm = re.search(r'R(\d{4})(\d{2})', org)
                if mm:
                    y, mo = int(mm.group(1)), int(mm.group(2))
                    if latest is None or (y, mo) > latest:
                        latest = (y, mo)
            return latest
        except Exception as e:
            self.logger.error(f'最新月探测失败: {e}')
            return None

    def discover_download_urls(self, year, month):
        """从文件索引中发现指定月份的Excel下载参数

        Args:
            year: 年份
            month: 月份

        Returns:
            dict or None: 下载参数字典
        """
        key = f'R{year}{month:02d}'
        pattern = re.compile(r'R' + str(year) + str(month).zfill(2) + r'(?!\d)')
        self.logger.info(f'发现 {year}-{month:02d} KAMA数据URL...')

        try:
            index = self._build_file_index()
            candidates = []
            for org, params in index.items():
                if pattern.search(org):
                    candidates.append(params)

            if not candidates:
                self.logger.warning(f'未找到 {key} 的下载参数')
                return None

            # 多个候选时优先 server_fileName 最新（修正版通常时间戳更大）
            candidates.sort(key=lambda p: p['server_fileName'])
            params = candidates[-1]
            self.logger.info(f'找到下载参数: {params["org_fileName"]} (候选 {len(candidates)} 个)')
            return params

        except Exception as e:
            self.logger.error(f'URL发现失败: {e}')
            return None

    # ------------------------------------------------------------
    # 下载
    # ------------------------------------------------------------
    def download_excel(self, params):
        """根据参数下载Excel文件

        Args:
            params: 下载参数字典

        Returns:
            bytes or None: Excel文件内容
        """
        if not params:
            return None

        try:
            resp = self.session.post(
                KAMA_FILE_DOWNLOAD_URL,
                data=params,
                headers=self.headers,
                timeout=60
            )

            if resp.status_code == 200 and len(resp.content) > 1000:
                self.logger.info(f'下载成功: {len(resp.content)} bytes')
                return resp.content
            else:
                self.logger.error(f'下载失败: HTTP {resp.status_code}, size={len(resp.content)}')
                return None

        except Exception as e:
            self.logger.error(f'下载异常: {e}')
            return None

    # ------------------------------------------------------------
    # 解析：仅使用 NR2 sheet（当月新登记，权威增量数据）
    # ------------------------------------------------------------
    def parse_new_registration_excel(self, excel_content, year, month):
        """解析新登记Excel数据（Sheet NR{YYYY}{MM}_2，当月增量）

        返回三条口径记录:
        1. 品牌级: 品牌합계行 col2
        2. 品牌×燃料级: 品牌합계行各燃料列
        3. 行业燃料级: 行业총계行各燃料列 (brand='KOREA INDUSTRY')
        """
        records = []

        try:
            wb = openpyxl.load_workbook(io.BytesIO(excel_content), data_only=True)

            # 查找 NR{MM}_2 Sheet
            sheet_name = None
            for name in wb.sheetnames:
                if 'NR' in name and name.endswith('_2'):
                    sheet_name = name
                    break

            if not sheet_name:
                self.logger.warning(f'未找到NR2 Sheet, 可用: {wb.sheetnames}')
                wb.close()
                return records

            self.logger.info(f'解析Sheet: {sheet_name}')
            ws = wb[sheet_name]

            # 定位表头行（구분|차종|합계|...）
            header_row = None
            for r in range(1, min(10, ws.max_row + 1)):
                v0 = ws.cell(r, 1).value
                v1 = ws.cell(r, 2).value
                v2 = ws.cell(r, 3).value
                if v0 == '구분' and v1 == '차종' and v2 == '합계':
                    header_row = r
                    break
            if header_row is None:
                self.logger.warning('未找到NR2表头行')
                wb.close()
                return records

            # 构建燃料列映射 (1-based列: 4..11)
            fuel_cols = {}
            for c in range(4, 12):
                v = ws.cell(header_row, c).value
                et = _map_fuel_type(v) if v is not None else None
                if et and et not in ('TOTAL',):
                    fuel_cols[c] = et
            self.logger.info(f'燃料列映射: {fuel_cols}')

            # 解析数据行
            current_brand = None
            for r in range(header_row + 1, ws.max_row + 1):
                gubun = ws.cell(r, 1).value
                cat = ws.cell(r, 2).value
                g = str(gubun).strip() if gubun is not None else ''
                cat_s = str(cat).strip() if cat is not None else ''

                if g == '총계':
                    # ============ 行业合计行 (col1 通常为空) ============
                    current_brand = 'KOREA INDUSTRY'
                    # 行业燃料级（忽略车型维度，避免唯一约束冲突）
                    for c, et in fuel_cols.items():
                        v = _to_int(ws.cell(r, c).value)
                        if v > 0:
                            records.append({
                                'country_code': 'KR',
                                'source_month': date(year, month, 1),
                                'brand_name_raw': 'KOREA INDUSTRY',
                                'brand_id': None,
                                'model_name': None,
                                'vehicle_type': 'ALL',
                                'energy_type': et,
                                'segment': None,
                                'raw_unit': 'units',
                                'sales_volume_raw': v,
                                'sales_volume_normalized': v,
                                'revision_no': 1,
                                'is_latest': True,
                                'pub_date': None,
                                'crawl_time': datetime.now(),
                                'data_source': 'kama',
                                'notes': 'KAMA 등록통계 신규등록 업계 연료별 (당월)',
                            })
                    continue

                # 更新当前品牌/板块
                if g and g not in ('합계', '계', '소계', 'TOTAL', 'SUBTOTAL'):
                    current_brand = g

                if not current_brand or cat_s != '합계':
                    continue

                # ============ 品牌合计行 ============
                total = _to_int(ws.cell(r, 3).value)
                if total > 0:
                    # 品牌级
                    records.append({
                        'country_code': 'KR',
                        'source_month': date(year, month, 1),
                        'brand_name_raw': current_brand,
                        'brand_id': None,
                        'model_name': None,
                        'vehicle_type': 'ALL',
                        'energy_type': None,
                        'segment': None,
                        'raw_unit': 'units',
                        'sales_volume_raw': total,
                        'sales_volume_normalized': total,
                        'revision_no': 1,
                        'is_latest': True,
                        'pub_date': None,
                        'crawl_time': datetime.now(),
                        'data_source': 'kama',
                        'notes': 'KAMA 등록통계 신규등록 제작사별 (당월)',
                    })
                    # 品牌×燃料级
                    for c, et in fuel_cols.items():
                        v = _to_int(ws.cell(r, c).value)
                        if v > 0:
                            records.append({
                                'country_code': 'KR',
                                'source_month': date(year, month, 1),
                                'brand_name_raw': current_brand,
                                'brand_id': None,
                                'model_name': None,
                                'vehicle_type': 'ALL',
                                'energy_type': et,
                                'segment': None,
                                'raw_unit': 'units',
                                'sales_volume_raw': v,
                                'sales_volume_normalized': v,
                                'revision_no': 1,
                                'is_latest': True,
                                'pub_date': None,
                                'crawl_time': datetime.now(),
                                'data_source': 'kama',
                                'notes': 'KAMA 등록통계 신규등록 제작사별 연료별 (당월)',
                            })

            wb.close()
            self.logger.info(f'解析到 {len(records)} 条新登记记录')
            return records

        except Exception as e:
            self.logger.error(f'解析新登记Excel失败: {e}')
            import traceback
            traceback.print_exc()

        return records

    # ------------------------------------------------------------
    # 爬取流程
    # ------------------------------------------------------------
    def crawl_month(self, year, month):
        """爬取指定月份的数据

        Returns:
            dict: 保存的记录数
        """
        self.logger.info(f'=== 爬取 {year}-{month:02d} KAMA数据 ===')

        params = self.discover_download_urls(year, month)
        if not params:
            self.logger.error(f'{year}-{month:02d} 下载参数未找到')
            return {'records': 0}

        excel_content = self.download_excel(params)
        if not excel_content:
            self.logger.error(f'{year}-{month:02d} 下载失败')
            return {'records': 0}

        records = self.parse_new_registration_excel(excel_content, year, month)

        saved = 0
        for record in records:
            if self.save_sales(record):
                saved += 1

        self.logger.info(f'{year}-{month:02d} 解析 {len(records)} 条, 保存 {saved} 条')
        return {'records': saved}

    def crawl_range(self, start_year, start_month, end_year, end_month):
        """爬取指定时间范围的数据

        Returns:
            dict: 各月份保存的记录数
        """
        results = {}

        current_year, current_month = start_year, start_month

        while (current_year, current_month) <= (end_year, end_month):
            key = f'{current_year}-{current_month:02d}'
            result = self.crawl_month(current_year, current_month)
            results[key] = result

            # 移动到下一月
            current_month += 1
            if current_month > 12:
                current_month = 1
                current_year += 1

        return results

    # ------------------------------------------------------------
    # 品牌匹配
    # ------------------------------------------------------------
    def get_brand_id(self, brand_name_raw):
        """获取品牌ID（支持韩语品牌名映射）"""
        brand_name = _map_korean_brand(brand_name_raw)

        try:
            conn = psycopg2.connect(**DB_CONFIG)
            cur = conn.cursor(cursor_factory=RealDictCursor)

            # 优先匹配canonical_name
            cur.execute("""
                SELECT id FROM brand_name_mapping
                WHERE LOWER(canonical_name) = LOWER(%s)
                  AND status = 'active'
                LIMIT 1
            """, (brand_name,))
            row = cur.fetchone()
            if row:
                cur.close()
                conn.close()
                return row['id']

            cur.close()
            conn.close()

        except Exception as e:
            self.logger.error(f'品牌ID查询失败: {e}')

        return None

    def save_sales(self, record):
        """重写保存方法，在保存前自动获取品牌ID"""
        record['brand_id'] = self.get_brand_id(record['brand_name_raw'])
        if record['brand_id']:
            self.logger.debug(f'品牌匹配成功: {record["brand_name_raw"]} -> {record["brand_id"]}')
        return super().save_sales(record)


def main():
    """主函数：爬取韩国KAMA数据 (2024-01 至 2026-07)"""
    crawler = KamaCrawler()

    print('=== KAMA韩国汽车销量爬虫 ===')
    print('数据口径: 신규등록 (当月新登记/注册量增量)')

    results = crawler.crawl_range(2024, 1, 2026, 7)

    for month_key, result in results.items():
        print(f'  {month_key}: 保存 {result["records"]} 条')

    print('\n完成！')


if __name__ == '__main__':
    main()
