# -*- coding: utf-8 -*-
"""
KBA 德国汽车销量爬虫
"""
import os
import sys
import json
import time
import random
import logging
import requests
import re
import io
from datetime import datetime, date
from bs4 import BeautifulSoup
from urllib.parse import urljoin

# 允许直接运行
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

# 从独立模块导入
from base_crawler import BaseCrawler, DB_CONFIG, UA_LIST


# ============================================
# KBA Crawler
# ============================================
class KBACrawler(BaseCrawler):
    def __init__(self):
        super().__init__('KBA', 'DE')
        # KBA 月度新车注册量产品目录页
        self.catalog_url = 'https://www.kba.de/DE/Statistik/Produktkatalog/produkte/Fahrzeuge/fz28/fz28_n_2026.html'
        self.base_download_url = 'https://www.kba.de/SharedDocs/Downloads/DE/Statistik/Fahrzeuge/FZ28/'
        self.session = requests.Session()
        self.session.headers.update(self.get_headers())

    def get_monthly_url(self, year, month):
        """构造某月数据文件地址，如 fz28_2024_01.xlsx（短月份2位）"""
        return f'{self.base_download_url}fz28_{year}_{month:02d}.xlsx?__blob=publicationFile'

    def latest_available_month(self):
        """探测 KBA 目录页最新可用月份，返回 (year, month) 或 None（不下载文件）"""
        try:
            resp = self.retry_request(self.session.get, self.catalog_url, timeout=30)
            if not resp:
                return None
            soup = BeautifulSoup(resp.text, 'html.parser')
            months = []
            for link in soup.find_all('a', href=True):
                href = link.get('href') or ''
                m = re.search(r'fz28_(\d{4})_(\d{2})', href)
                if m:
                    months.append((int(m.group(1)), int(m.group(2))))
            if not months:
                return None
            return max(months)
        except Exception as e:
            self.logger.warning(f'KBA 最新月探测失败: {e}')
            return None

    def crawl_latest(self):
        self.logger.info('开始爬取 KBA 最新月度数据')
        self.random_delay()

        # 1. 优先尝试从目录页自动发现最新 XLSX
        latest_xlsx = None
        try:
            resp = self.retry_request(self.session.get, self.catalog_url, timeout=30)
            if resp:
                soup = BeautifulSoup(resp.text, 'html.parser')
                xlsx_links = []
                for link in soup.find_all('a', href=True):
                    href = link.get('href') or ''
                    if href and 'fz28' in href.lower():
                        full_url = urljoin(self.catalog_url, href)
                        xlsx_links.append(full_url)

                if xlsx_links:
                    latest_xlsx = sorted(xlsx_links)[-1]
                    self.logger.info(f'从目录页找到最新 XLSX: {latest_xlsx}')
        except Exception as e:
            self.logger.warning(f'目录页解析失败: {e}')

        # 2. 回退：硬编码最新 XLSX（2026年6月）
        if not latest_xlsx:
            latest_xlsx = 'https://www.kba.de/SharedDocs/Downloads/DE/Statistik/Fahrzeuge/FZ28/fz28_2026_06.xlsx?__blob=publicationFile&v=2'
            self.logger.info(f'使用回退地址: {latest_xlsx}')

        # 2. 下载 XLSX
        xlsx_content = self.retry_request(self.session.get, latest_xlsx, timeout=60)
        if not xlsx_content:
            return False

        # 3. 解析 XLSX
        try:
            # 从文件名推导数据月份，如 fz28_2026_06.xlsx → 2026-06
            source_month = self._extract_source_month(latest_xlsx)
            self.logger.info(f'解析到数据月份: {source_month}')

            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(xlsx_content.content), data_only=True)
            
            # 优先使用 FZ 28.4（按品牌划分的 Pkw 销量表）
            if 'FZ 28.4' in wb.sheetnames:
                ws = wb['FZ 28.4']
            else:
                ws = wb.active
            
            tables = []
            for row in ws.iter_rows():
                tables.append([cell.value for cell in row])

            records = self._parse_xlsx_tables(tables, source_month)
            self.logger.info(f'解析到 {len(records)} 条记录')

            saved = 0
            for record in records:
                if self.save_sales(record):
                    saved += 1

            self.logger.info(f'保存成功 {saved}/{len(records)} 条')
            self.page_count += 1
            self.batch_restart()
            return saved > 0

        except ImportError:
            self.logger.error('请安装 openpyxl: pip install openpyxl')
            return False
        except Exception as e:
            self.logger.error(f'XLSX 解析失败: {e}')
            return False

    def _extract_source_month(self, url):
        """从下载地址/文件名推导数据月份，如 fz28_2026_06.xlsx → 2026-06-01"""
        # 兼容两种命名：fz28_2026_06.xlsx 与 fz28_2026_202606.xlsx
        m = re.search(r'(\d{4})_(\d{2})\.xlsx', url)
        if m:
            return date(int(m.group(1)), int(m.group(2)), 1)
        m2 = re.search(r'(\d{4})_(\d{6})\.xlsx', url)
        if m2:
            return date(int(m2.group(1)), int(m2.group(2)[4:6]), 1)
        return datetime.now().replace(day=1).date()

    def _parse_xlsx_tables(self, tables, source_month=None, is_latest=True):
        records = []
        if source_month is None:
            source_month = datetime.now().replace(day=1).date()

        # tables 是一个 list，每个元素是一行（list of cell values）
        # 找到包含 'Marke' 的表头行索引
        header_idx = None
        for idx, row in enumerate(tables):
            if row and any('Marke' in str(cell) for cell in row if cell):
                header_idx = idx
                break
        
        if header_idx is None:
            self.logger.warning('未找到 Marke 表头行')
            return records
        
        self.logger.info(f'表头行索引: {header_idx}')
        
        # 从表头后开始解析品牌数据行
        for row in tables[header_idx + 1:]:
            if not row or len(row) < 5:
                continue
            
            brand_raw = str(row[1]).strip() if row[1] else ''
            volume_raw = str(row[2]).strip() if row[2] else ''
            share_raw = str(row[4]).strip() if row[4] else ''
            
            if not brand_raw or not volume_raw or brand_raw == 'INSGESAMT':
                continue
            # 跳过"其他品牌"汇总行（SONSTIGE），其数值是杂牌总和，非品牌数据
            if brand_raw.upper() == 'SONSTIGE':
                continue
            
            try:
                volume = int(re.sub(r'[^\d]', '', volume_raw))
            except Exception:
                continue
            
            share = None
            if share_raw and share_raw != '-':
                try:
                    share = float(re.sub(r'[^\d.]', '', share_raw))
                except Exception:
                    pass
            
            brand_id = self._match_brand(brand_raw)
            
            record = {
                'country_code': 'DE',
                'source_month': source_month,
                'brand_name_raw': brand_raw,
                'brand_id': brand_id,
                'model_name': None,
                'vehicle_type': 'passenger',
                'energy_type': 'unknown',
                'segment': None,
                'raw_unit': 'unit',
                'sales_volume_raw': volume,
                'sales_volume_normalized': volume,
                'revision_no': 1,
                'is_latest': is_latest,
                'pub_date': datetime.now().date(),
                'crawl_time': datetime.now(),
                'data_source': 'KBA',
                'notes': f'KBA monthly {datetime.now().strftime("%Y-%m")}'
            }
            records.append(record)
        
        return records

    def _detect_energy_type(self, text, is_summary_row=False):
        """从行文本中识别能源类型，仅在汇总行(is_summary_row=True)时触发切换"""
        if not is_summary_row:
            return None
        t = text.lower()
        # 跳过alt大类（Alternativer Antrieb），只保留具体能源类型
        if 'alternativer antrieb' in t:
            return None
        if 'elektro (bev)' in t or 'elektro(bev)' in t:
            return 'bev'
        elif 'plug-in-hybrid' in t or 'plug.in.hybrid' in t:
            return 'phev'
        elif 'voll-hybrid' in t or 'voll.hybrid' in t:
            return 'hev'
        elif 'brennstoffzelle' in t:
            return 'fcev'
        elif 'gas' in t:
            return 'gas'
        elif 'wasserstoff' in t:
            return 'hydrogen'
        elif 'elektro-antriebe' in t or 'elektro antriebe' in t:
            return 'bev'
        return None

    def _parse_model_tables(self, ws, source_month=None, is_latest=False):
        """解析 FZ 28.7 车型级数据 Sheet"""
        records = []
        if source_month is None:
            source_month = datetime.now().replace(day=1).date()

        SKIP_KEYWORDS = ['insgesamt', 'sonstige', 'anteil', 'anzahl']
        current_energy_type = None

        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True)):
            vals = [str(c).strip() if c is not None else '' for c in row[:7]]
            if not any(vals):
                continue

            model_raw = vals[1] if len(vals) > 1 else ''
            lower_model = model_raw.lower()
            row_text = ' '.join(vals)

            # 判断是否为汇总/跳过行
            is_skip = i < 10 or any(k in lower_model for k in SKIP_KEYWORDS)

            # 仅在汇总行上识别能源类型（避免车型行误触发）
            if is_skip:
                detected = self._detect_energy_type(row_text, is_summary_row=True)
                if detected:
                    current_energy_type = detected
                    self.logger.info(f'行{i+1}: 能源类型切换为 {detected}')

            # 跳过表头/目录/汇总行
            if is_skip:
                continue

            # 未进入具体能源分类前，跳过车型数据（alt大类不存储）
            if current_energy_type is None:
                continue

            # 车型名格式为 "BRAND MODEL"，按第一个空格拆分
            if ' ' in model_raw and len(model_raw) > 5:
                parts = model_raw.split(' ', 1)
                brand_name_raw = parts[0]
                model_name = parts[1]

                # 提取当月销量 (Col C)
                try:
                    volume = int(float(vals[2].replace('.', '')))
                except (ValueError, IndexError):
                    volume = 0

                if volume > 0:
                    brand_id = self._match_brand(brand_name_raw)

                    record = {
                        'country_code': 'DE',
                        'source_month': source_month,
                        'brand_name_raw': brand_name_raw,
                        'brand_id': brand_id,
                        'model_name': model_name,
                        'vehicle_type': 'passenger',
                        'energy_type': current_energy_type or 'unknown',
                        'segment': None,
                        'raw_unit': 'unit',
                        'sales_volume_raw': volume,
                        'sales_volume_normalized': volume,
                        'revision_no': 1,
                        'is_latest': is_latest,
                        'pub_date': datetime.now().date(),
                        'crawl_time': datetime.now(),
                        'data_source': 'KBA',
                        'notes': f'KBA model-level {datetime.now().strftime("%Y-%m")}'
                    }
                    records.append(record)

        self.logger.info(f'FZ 28.7 车型级解析完成: {len(records)} 条')
        return records

    def crawl_one_month(self, year, month, include_model=False):
        """
        下载解析并入库某个月的数据
        include_model: 是否同时解析车型级数据（FZ 28.7）
        """
        url = self.get_monthly_url(year, month)
        source_month = date(year, month, 1)
        self.logger.info(f'下载 {url}')
        resp = self.retry_request(self.session.get, url, timeout=60)
        if not resp or resp.status_code != 200:
            self.logger.error(f'{year}-{month:02d} 下载失败: HTTP {resp.status_code if resp else "None"}')
            return 0, 0
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(resp.content), data_only=True)

            brand_saved = 0
            model_saved = 0

            # 1. 品牌级解析（FZ 28.4）
            if 'FZ 28.4' in wb.sheetnames:
                ws_brand = wb['FZ 28.4']
                tables = [[cell.value for cell in row] for row in ws_brand.iter_rows()]
                brand_records = self._parse_xlsx_tables(tables, source_month, is_latest=False)
                brand_saved = sum(1 for r in brand_records if self.save_sales(r))
                self.logger.info(f'{year}-{month:02d} 品牌级: 解析 {len(brand_records)} 条，入库 {brand_saved} 条')

            # 2. 车型级解析（FZ 28.7，可选）
            if include_model and 'FZ 28.7' in wb.sheetnames:
                ws_model = wb['FZ 28.7']
                model_records = self._parse_model_tables(ws_model, source_month, is_latest=False)
                model_saved = sum(1 for r in model_records if self.save_sales(r))
                self.logger.info(f'{year}-{month:02d} 车型级: 解析 {len(model_records)} 条，入库 {model_saved} 条')

            self.page_count += 1
            self.batch_restart()
            return brand_saved, model_saved
        except Exception as e:
            self.logger.error(f'{year}-{month:02d} 解析失败: {e}')
            return 0, 0

    def crawl_range(self, start_year, start_month, end_year, end_month, include_model=False):
        """
        历史回放：按月遍历从(start_year,start_month)到(end_year,end_month)的数据。
        include_model: 是否同时解析车型级数据（FZ 28.7）
        """
        self.logger.info(f'KBA 历史回放 {start_year}-{start_month:02d} ~ {end_year}-{end_month:02d} (含车型级={include_model})')
        cursor_y, cursor_m = start_year, start_month
        total_brand_saved = 0
        total_model_saved = 0
        last_brand_success = None
        last_model_success = None

        while (cursor_y, cursor_m) <= (end_year, end_month):
            brand_saved, model_saved = self.crawl_one_month(cursor_y, cursor_m, include_model=include_model)
            if brand_saved > 0:
                last_brand_success = (cursor_y, cursor_m)
            if model_saved > 0:
                last_model_success = (cursor_y, cursor_m)
            total_brand_saved += brand_saved
            total_model_saved += model_saved
            self.random_delay()
            cursor_m += 1
            if cursor_m > 12:
                cursor_m = 1
                cursor_y += 1

        # 标记品牌级最新月
        if last_brand_success:
            conn, cur = self.get_connection()
            ly, lm = last_brand_success
            cur.execute("""
                UPDATE market_sales_monthly SET is_latest = FALSE
                WHERE country_code = 'DE' AND data_source = 'KBA' AND model_name IS NULL
            """)
            cur.execute("""
                UPDATE market_sales_monthly SET is_latest = TRUE
                WHERE country_code = 'DE' AND data_source = 'KBA'
                  AND source_month = %s AND model_name IS NULL
            """, (date(ly, lm, 1),))
            conn.commit()
            self.logger.info(f'品牌级最后成功月 {ly}-{lm:02d} 标记为最新')

        # 标记车型级最新月
        if include_model and last_model_success:
            conn, cur = self.get_connection()
            ly, lm = last_model_success
            cur.execute("""
                UPDATE market_sales_monthly SET is_latest = FALSE
                WHERE country_code = 'DE' AND data_source = 'KBA' AND model_name IS NOT NULL
            """)
            cur.execute("""
                UPDATE market_sales_monthly SET is_latest = TRUE
                WHERE country_code = 'DE' AND data_source = 'KBA'
                  AND source_month = %s AND model_name IS NOT NULL
            """, (date(ly, lm, 1),))
            conn.commit()
            self.logger.info(f'车型级最后成功月 {ly}-{lm:02d} 标记为最新')

        self.logger.info(f'历史回放完成：品牌级 {total_brand_saved} 条，车型级 {total_model_saved} 条')
        self.close()
        return total_brand_saved, total_model_saved

    def run(self, months=1):
        self.logger.info(f'KBA 爬虫启动，目标 {months} 个月')
        success_count = 0

        for i in range(months):
            self.logger.info(f'爬取第 {i+1}/{months} 个月')
            if self.crawl_latest():
                success_count += 1

        self.logger.info(f'完成：成功 {success_count}/{months} 个月')
        self.close()
        return success_count


if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser(description='KBA 德国销量爬虫')
    parser.add_argument('--range', action='store_true', help='历史回放模式（默认2024-01至今）')
    parser.add_argument('--model', action='store_true', help='同时爬取车型级数据（FZ 28.7）')
    parser.add_argument('--start', default='2024-01', help='起始年月，如 2024-01')
    parser.add_argument('--end', default=None, help='结束年月，如 2026-06，默认至今')
    args = parser.parse_args()

    crawler = KBACrawler()
    if args.range:
        sy, sm = map(int, args.start.split('-'))
        ey, em = 2026, 6
        if args.end:
            ey, em = map(int, args.end.split('-'))
        crawler.crawl_range(sy, sm, ey, em, include_model=args.model)
    else:
        crawler.run(months=1)
