# -*- coding: utf-8 -*-
"""ACAP 葡萄牙汽车月度销量爬虫"""
import sys
sys.path.insert(0, '.')
import re
import io
import logging
import requests
import psycopg2
import openpyxl
from psycopg2.extras import RealDictCursor
from datetime import datetime, date
from bs4 import BeautifulSoup
from base_crawler import BaseCrawler, DB_CONFIG

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s | %(levelname)s | %(message)s'
)

PT_BASE_URL = 'https://www.acap.pt'
PT_DADOS_URL = 'https://www.acap.pt/pt/estatisticas/dados'

PT_MONTH_MAP = {
    'janeiro': 1, 'fevereiro': 2, 'marco': 3, 'março': 3,
    'abril': 4, 'maio': 5, 'junho': 6, 'julho': 7,
    'agosto': 8, 'setembro': 9, 'outubro': 10, 'novembro': 11, 'dezembro': 12,
}

# 注意：market_sales_monthly 唯一约束不含 vehicle_type，同月同品牌不同车型会互相覆盖，
# 因此只入库乘用车（Lig. Passageiros）品牌级，与其他国家口径保持一致。
VEHICLE_TYPE_MAP = {
    'Lig. Passageiros': 'passenger',
}

SKIP_BRANDS = ('TOTAL', 'TOTAL GERAL', 'OUTROS', 'OUTRAS')

# 经销商后缀（如 "Volkswagen - Auto Ribeiro" → "Volkswagen"）
DEALER_SUFFIX_RE = re.compile(r'\s*-\s*[^\-]+$')


def _to_int(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return int(value)
    s = str(value).strip()
    s = s.replace('\xa0', '').replace(' ', '').replace(',', '')
    s = s.replace('.', '')
    s = re.sub(r'[^\d]', '', s)
    return int(s) if s else None


def _normalize_brand(brand_raw):
    """归一化品牌名：剥离经销商后缀（- Auto Ribeiro 等）、去重音符号（Citroën→Citroen）"""
    s = str(brand_raw).strip()
    s = DEALER_SUFFIX_RE.sub('', s).strip()
    import unicodedata
    s = ''.join(c for c in unicodedata.normalize('NFKD', s) if not unicodedata.combining(c))
    return s


class PtCrawler(BaseCrawler):
    """ACAP 葡萄牙汽车销量爬虫"""

    def __init__(self):
        super().__init__(source_name='acap', country_code='PT')
        self.headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Referer': PT_DADOS_URL,
        }
        self.session = requests.Session()

    def discover_files(self):
        """从数据页发现品牌级xlsx和能源PDF的URL"""
        resp = self.session.get(PT_DADOS_URL, headers=self.headers, timeout=30)
        resp.encoding = resp.apparent_encoding
        soup = BeautifulSoup(resp.text, 'html.parser')
        xlsx_url = None
        energy_url = None
        for a in soup.find_all('a', href=True):
            href = a['href']
            txt = a.get_text(strip=True)
            if 'xlsx' in href and 'Marca' in txt and not xlsx_url:
                xlsx_url = href
            if 'pdf' in href and 'energia' in txt and not energy_url:
                energy_url = href
        if xlsx_url and not xlsx_url.startswith('http'):
            xlsx_url = PT_BASE_URL + xlsx_url
        if energy_url and not energy_url.startswith('http'):
            energy_url = PT_BASE_URL + energy_url
        self.logger.info(f'发现品牌xlsx: {xlsx_url}')
        self.logger.info(f'发现能源pdf: {energy_url}')
        return xlsx_url, energy_url

    def download_excel(self, url):
        resp = self.session.get(url, headers=self.headers, timeout=60)
        if resp.status_code == 200 and len(resp.content) > 1000:
            self.logger.info(f'下载成功: {len(resp.content)} bytes')
            return resp.content
        self.logger.error(f'下载失败: HTTP {resp.status_code}')
        return None

    def parse_excel(self, excel_content, target_month):
        """解析品牌级xlsx，只取Lig. Passageiros（乘用车）sheet

        返回 (records, found_month_name)
        """
        records = []
        found_month = None
        try:
            wb = openpyxl.load_workbook(io.BytesIO(excel_content), data_only=True)
            for sheet_name in wb.sheetnames:
                if sheet_name not in VEHICLE_TYPE_MAP:
                    continue
                ws = wb[sheet_name]
                # 定位表头：找含 'Unidades' 的行（r7或r6）
                header_row_idx = None
                month_name = None
                month = None
                for i in range(1, min(ws.max_row, 15) + 1):
                    row_vals = [ws.cell(row=i, column=c).value for c in range(1, 13)]
                    if any(v is not None and 'unidades' in str(v).lower() for v in row_vals):
                        header_row_idx = i
                        # 月份名在上一行
                        prev_vals = [ws.cell(row=i - 1, column=c).value for c in range(1, 13)]
                        for v in prev_vals:
                            if v is not None and str(v).strip().lower() in PT_MONTH_MAP:
                                month_name = str(v).strip()
                                month = PT_MONTH_MAP[month_name.lower()]
                                break
                        # 年份在表头下一行（r8或r7）
                        yr_vals = [ws.cell(row=i + 1, column=c).value for c in range(1, 13)]
                        year = None
                        for v in yr_vals:
                            if v is not None and re.fullmatch(r'20\d\d', str(v).strip()):
                                year = int(str(v).strip())
                                break
                        break
                if header_row_idx is None:
                    self.logger.warning(f'{sheet_name}: 未找到表头')
                    continue
                if month is None:
                    self.logger.warning(f'{sheet_name}: 未找到月份名')
                    continue
                if found_month is None:
                    found_month = (year, month)
                self.logger.info(f'{sheet_name}: 表头行={header_row_idx} 月份={month_name} {year}')

                # 数据从 header_row_idx+2 开始（r9或r8），品牌=col1，当月=col2
                data_start = header_row_idx + 2
                for r in range(data_start, ws.max_row + 1):
                    brand = ws.cell(row=r, column=1).value
                    if brand is None:
                        continue
                    brand_str = _normalize_brand(brand)
                    if not brand_str:
                        continue
                    if brand_str.upper() in SKIP_BRANDS or brand_str.upper().startswith('TOTAL'):
                        continue
                    qty = _to_int(ws.cell(row=r, column=2).value)
                    if qty is None or qty <= 0:
                        continue
                    records.append({
                        'country_code': 'PT',
                        'source_month': date(year, month, 1),
                        'brand_name_raw': brand_str,
                        'brand_id': None,
                        'model_name': None,
                        'vehicle_type': VEHICLE_TYPE_MAP[sheet_name],
                        'energy_type': None,
                        'segment': None,
                        'raw_unit': 'units',
                        'sales_volume_raw': qty,
                        'sales_volume_normalized': qty,
                        'revision_no': 1,
                        'is_latest': True,
                        'pub_date': None,
                        'crawl_time': datetime.now(),
                        'data_source': 'acap',
                        'notes': f'ACAP matrículas {sheet_name} {year}-{month:02d}',
                    })
            wb.close()
            self.logger.info(f'解析到 {len(records)} 条记录')
        except Exception as e:
            self.logger.error(f'解析失败: {e}')
            import traceback
            traceback.print_exc()
        return records, found_month

    def crawl_month(self, year, month):
        """爬取指定月份（xlsx为滚动快照，需与数据页当前月份一致）"""
        self.logger.info(f'=== 爬取 {year}-{month:02d} ACAP数据 ===')
        xlsx_url, energy_url = self.discover_files()
        if not xlsx_url:
            self.logger.error('未找到品牌xlsx')
            return {'records': 0}
        excel_content = self.download_excel(xlsx_url)
        if not excel_content:
            return {'records': 0}
        records, found_month = self.parse_excel(excel_content, (year, month))
        if found_month and found_month != (year, month):
            self.logger.warning(f'数据页月份 {found_month} 与目标 {year}-{month:02d} 不符，跳过入库')
            return {'records': 0}
        saved = 0
        for record in records:
            if self.save_sales(record):
                saved += 1
        self.logger.info(f'{year}-{month:02d} 保存 {saved} 条')
        return {'records': saved}

    def latest_available_month(self):
        """从xlsx解析数据页当前月份"""
        xlsx_url, _ = self.discover_files()
        if not xlsx_url:
            return None
        excel_content = self.download_excel(xlsx_url)
        if not excel_content:
            return None
        _, found_month = self.parse_excel(excel_content, None)
        return found_month

    def crawl_incremental(self):
        """增量：解析数据页当前月份，若>库中MAX则入库"""
        conn, cur = self.get_connection()
        cur.execute("SELECT MAX(source_month) AS m FROM market_sales_monthly WHERE country_code='PT' AND data_source='acap'")
        row = cur.fetchone()
        max_month = row['m'] if row else None
        xlsx_url, _ = self.discover_files()
        if not xlsx_url:
            return 0
        excel_content = self.download_excel(xlsx_url)
        if not excel_content:
            return 0
        records, found_month = self.parse_excel(excel_content, None)
        if not found_month:
            return 0
        latest = date(found_month[0], found_month[1], 1)
        if max_month and latest <= max_month:
            self.logger.info(f'PT 数据页 {latest} <= 库中MAX {max_month}, 无新数据')
            return 0
        saved = 0
        for record in records:
            if self.save_sales(record):
                saved += 1
        self.logger.info(f'PT 增量保存 {saved} 条 ({latest})')
        return saved

    def get_brand_id(self, brand_name_raw):
        brand_upper = brand_name_raw.strip().upper()
        try:
            conn = psycopg2.connect(**DB_CONFIG)
            cur = conn.cursor(cursor_factory=RealDictCursor)
            cur.execute("SELECT id FROM brand_name_mapping WHERE UPPER(canonical_name)=%s OR UPPER(brand_name_cn)=%s LIMIT 1", (brand_upper, brand_upper))
            row = cur.fetchone()
            if not row:
                cur.execute("SELECT brand_id FROM brand_name_variant WHERE UPPER(variant_name)=%s LIMIT 1", (brand_upper,))
                row = cur.fetchone()
            cur.close()
            conn.close()
            return row['id'] if row else None
        except Exception as e:
            self.logger.error(f'品牌ID查询失败: {e}')
            return None

    def save_sales(self, record):
        record['brand_id'] = self.get_brand_id(record['brand_name_raw'])
        return super().save_sales(record)


def main():
    crawler = PtCrawler()
    crawler.crawl_incremental()


if __name__ == '__main__':
    main()
