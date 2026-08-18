# -*- coding: utf-8 -*-
"""
CH 瑞士汽车月度销量爬虫
数据源: https://www.auto.swiss/ (auto-schweiz → 301 → auto.swiss)
数据类型: 品牌级乘用车新登记 (Neuzulassungen, Personenwagen)
数据来源: auto-schweiz ASTRA/OFROU / MOFIS

Excel结构:
- 2021-2026: MOFISPW{YYYY}_{M}.xlsx (当年滚动, 每sheet一个月份)
- 2015-2020: auto-schweiz_PW_{YYYY}[_12].xlsx
- 2011-2014: auto-schweiz_PW_{YYYY}.xls
每个sheet: 行9/10=表头(Marken/marques), 品牌行从表头+1, 数量列=索引3(列D), Total行后为燃料类别

下载机制:
- GET https://www.auto.swiss/ 解析首页提取Excel链接
- GET https://www.auto.swiss/wp-content/uploads/.../xxx.xlsx
"""
import sys
sys.path.insert(0, '.')
import os
import re
import io
import logging
import requests
import openpyxl
from datetime import datetime, date
from urllib.parse import urljoin
from bs4 import BeautifulSoup
from base_crawler import BaseCrawler, DB_CONFIG

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s | %(levelname)s | %(message)s'
)

CH_HOME_URL = 'https://www.auto.swiss/'

# sheet名 -> 月份
SHEET_MONTH_MAP = {
    'jan.': 1, 'jan': 1, 'januar': 1,
    'feb.': 2, 'feb': 2, 'februar': 2,
    'mär.': 3, 'märz': 3, 'mrz': 3, 'marz': 3, 'mar': 3,
    'apr.': 4, 'april': 4, 'apr': 4,
    'mai.': 5, 'mai': 5, 'may': 5,
    'jun.': 6, 'juni': 6, 'jun': 6,
    'jul.': 7, 'juli': 7, 'jul': 7,
    'aug.': 8, 'aug': 8, 'august': 8,
    'sept.': 9, 'sept': 9, 'sep.': 9, 'sep': 9, 'september': 9,
    'okt.': 10, 'okt': 10, 'oktober': 10,
    'nov.': 11, 'nov': 11, 'november': 11,
    'dez.': 12, 'dez': 12, 'dezember': 12, 'dec': 12, 'december': 12,
}

# 跳过行 (品牌起始行后)
SKIP_ROWS = ('TOTAL', 'GESAMT', 'DIVERSE MARKEN', 'QUELLE', 'DAVON',
             '4 X 4', 'BENZIN', 'DIESEL', 'HYBRID', 'PLUG-IN', 'ELEKTRISCH',
             'CNG', 'WASSERSTOFF', 'SUMME ÜBRIGE', 'SUMME ÜBRIGER', 'STICHTAG',
             'MARKEN / MARQUES')

# 瑞士官方合并行/重音品牌 → 归一化后品牌名 (映射到 brand_name_mapping 可匹配形式)
BRAND_MERGE_MAP = {
    'SEAT / CUPRA': 'SEAT',        # 瑞士官方合并行, 归 SEAT(id=13)
    'KGM / SSANGYONG': 'KGM',      # 合并行, 归 KGM(id=215)
    'SKODA': 'SKODA',              # 占位, 重音由 NFKD 归一化处理 (Š→S)
    'CITROEN': 'CITROEN',
}

def _normalize_brand(brand_raw):
    """归一化瑞士品牌名: 去重音(Š→S/ë→e) + 合并行处理 + 大小写统一"""
    import unicodedata
    brand = str(brand_raw).strip()
    upper = brand.upper()
    # 合并行优先处理
    if ' / ' in brand:
        parts = [p.strip() for p in brand.split('/')]
        if len(parts) == 2:
            merged = f'{parts[0]} / {parts[1]}'.upper()
            mapped = BRAND_MERGE_MAP.get(merged)
            if mapped:
                return mapped
    # 去重音符号 (Š→S, ë→e, ä→a ...)
    brand = unicodedata.normalize('NFKD', brand)
    brand = ''.join(c for c in brand if not unicodedata.combining(c))
    return brand


class SwitzerlandCrawler(BaseCrawler):
    """瑞士auto.swiss汽车月度销量爬虫"""

    def __init__(self):
        super().__init__(source_name='auto-schweiz', country_code='CH')
        self.headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Referer': CH_HOME_URL,
        }
        self.session = requests.Session()

    def discover_excel_links(self):
        """从首页提取品牌级Excel下载链接
        
        Returns:
            list of {year: int, url: str, filename: str}
        """
        links = []
        try:
            resp = self.session.get(CH_HOME_URL, headers=self.headers, timeout=30)
            if resp.status_code != 200:
                self.logger.error(f'首页访问失败: HTTP {resp.status_code}')
                return links
            resp.encoding = 'utf-8'
            soup = BeautifulSoup(resp.text, 'html.parser')

            for a in soup.find_all('a', href=True):
                href = a['href']
                # 只取品牌级文件 (MOFISPW 或 PW, 排除 Modelle/Stecker/NFMOFIS/FahrzeugemitStecker)
                if not re.search(r'(MOFISPW|PW_\d{4})', href):
                    continue
                if re.search(r'(Modelle|Stecker|NFMOFIS|Fahrzeugemit)', href, re.IGNORECASE):
                    continue
                filename = href.split('/')[-1]
                m = re.search(r'MOFISPW(\d{4})', filename) or re.search(r'PW_(\d{4})', filename)
                if not m:
                    continue
                year = int(m.group(1))
                if not (2010 <= year <= 2026):
                    continue
                full_url = urljoin(CH_HOME_URL, href)
                links.append({'year': year, 'url': full_url, 'filename': filename})

            self.logger.info(f'发现 {len(links)} 个品牌级Excel链接')
            for l in sorted(links, key=lambda x: x['year']):
                self.logger.info(f"  {l['year']}: {l['filename']}")
            return links
        except Exception as e:
            self.logger.error(f'链接发现失败: {e}')
            return links

    def download_excel(self, url):
        """下载Excel文件"""
        try:
            resp = self.session.get(url, headers=self.headers, timeout=60)
            if resp.status_code == 200 and len(resp.content) > 1000:
                self.logger.info(f'下载成功: {len(resp.content)} bytes')
                return resp.content
            else:
                self.logger.error(f'下载失败: HTTP {resp.status_code}, size={len(resp.content)}')
                return None
        except Exception as e:
            self.logger.error(f'下载异常: {e}')
            return None

    def _to_int(self, value):
        try:
            if value is None:
                return None
            return int(round(float(value)))
        except (ValueError, TypeError):
            return None

    def parse_file(self, excel_content, year):
        """解析一个Excel文件的所有月份sheet

        支持 .xlsx (openpyxl) 和 .xls (pandas+xlrd)

        Returns:
            list of records
        """
        records = []
        try:
            import pandas as pd
            if excel_content[:8] == b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1':
                # 旧 .xls (OLE格式): sheet -> DataFrame 转二维列表
                sheets_raw = pd.read_excel(io.BytesIO(excel_content), sheet_name=None, header=None)
                sheets_data = {sn: df.values.tolist() for sn, df in sheets_raw.items()}
            else:
                # 新 .xlsx (zip格式)
                wb = openpyxl.load_workbook(io.BytesIO(excel_content), data_only=True)
                sheets_data = {}
                for sn in wb.sheetnames:
                    rows = [[c.value for c in row] for row in wb[sn].iter_rows()]
                    sheets_data[sn] = rows
                wb.close()

            for sheet_name, rows in sheets_data.items():
                month = SHEET_MONTH_MAP.get(str(sheet_name).lower().strip())
                if not month:
                    self.logger.debug(f'跳过非月份sheet: {sheet_name}')
                    continue

                header_row = None
                for row_idx, row in enumerate(rows[:15]):
                    if row and row[0] and isinstance(row[0], str) and 'Marken' in row[0]:
                        header_row = row_idx
                        break
                if header_row is None:
                    self.logger.warning(f'{year}-{month:02d} 未找到表头, sheets={sheet_name}')
                    continue

                start = header_row + 1
                for row_idx in range(start, len(rows)):
                    row = rows[row_idx]
                    if not row or not row[0]:
                        continue
                    brand_raw = row[0]
                    sales = self._to_int(row[3]) if len(row) > 3 else None
                    brand = str(brand_raw).strip()
                    if not brand:
                        continue
                    # 遇到 Total/GESAMT/Diverse/燃料类别/Quelle 行则停止本sheet
                    if brand.upper().startswith('TOTAL') or brand.upper().startswith('GESAMT') \
                            or 'Diverse' in brand or brand.upper().startswith('DAVON') \
                            or brand.upper().startswith('QUELLE'):
                        break
                    if sales is None or sales <= 0:
                        continue

                    # 归一化品牌名 (重音/合并行)
                    norm_brand = _normalize_brand(brand)

                    records.append({
                        'country_code': 'CH',
                        'source_month': date(year, month, 1),
                        'brand_name_raw': norm_brand,
                        'brand_id': None,
                        'model_name': None,
                        'vehicle_type': 'passenger_car',
                        'energy_type': None,
                        'segment': None,
                        'raw_unit': 'units',
                        'sales_volume_raw': sales,
                        'sales_volume_normalized': sales,
                        'revision_no': 1,
                        'is_latest': True,
                        'pub_date': None,
                        'crawl_time': datetime.now(),
                        'data_source': 'auto-schweiz',
                        'notes': f'auto-schweiz PW Neuzulassungen {year}-{month:02d} (CH & FL)',
                    })
            self.logger.info(f'{year} 文件解析到 {len(records)} 条记录')
        except Exception as e:
            self.logger.error(f'解析Excel失败({year}): {e}')
            import traceback
            traceback.print_exc()
        return records

    def crawl_incremental(self):
        """增量更新：发现所有年度Excel，解析后只保存比库中MAX(source_month)更新的月份"""
        links = self.discover_excel_links()
        if not links:
            self.logger.error('未发现任何Excel链接')
            return {}

        conn, cur = self.get_connection()
        cur.execute("SELECT MAX(source_month) AS m FROM market_sales_monthly WHERE country_code='CH'")
        row = cur.fetchone()
        max_month = row['m'] if row else None
        self.logger.info(f'CH 增量: 库中已有最大月 {max_month}')

        results = {}
        for link in sorted(links, key=lambda x: x['year']):
            year = link['year']
            if max_month is not None and date(year, 12, 1) <= max_month:
                results[year] = {'records': 0, 'saved': 0, 'skipped': 'already_up_to_date'}
                continue
            self.logger.info(f'=== 处理 {year} 数据 ===')
            content = self.download_excel(link['url'])
            if not content:
                results[year] = {'records': 0, 'error': 'download_fail'}
                continue
            records = self.parse_file(content, year)
            new_records = [r for r in records if max_month is None or r['source_month'] > max_month]
            saved = 0
            for rec in new_records:
                if self.save_sales(rec):
                    saved += 1
            results[year] = {'records': len(records), 'saved': saved, 'new': len(new_records)}
            self.logger.info(f'{year}: 解析{len(records)}条, 新{len(new_records)}条, 保存{saved}条')
        return results

    def crawl_all(self):
        """爬取全部年度数据"""
        links = self.discover_excel_links()
        if not links:
            self.logger.error('未发现任何Excel链接')
            return {}

        results = {}
        # 按年份升序处理
        for link in sorted(links, key=lambda x: x['year']):
            year = link['year']
            self.logger.info(f'=== 爬取 {year} 数据 ===')
            content = self.download_excel(link['url'])
            if not content:
                results[year] = {'records': 0, 'error': 'download_fail'}
                continue
            records = self.parse_file(content, year)
            saved = 0
            for rec in records:
                if self.save_sales(rec):
                    saved += 1
            results[year] = {'records': len(records), 'saved': saved}
            self.logger.info(f'{year}: 解析{len(records)}条, 保存{saved}条')
        return results


def main():
    crawler = SwitzerlandCrawler()
    print('=== CH 瑞士 auto.swiss 爬虫 ===')
    results = crawler.crawl_all()
    for year, res in sorted(results.items()):
        print(f'  {year}: 解析 {res["records"]} 条, 保存 {res["saved"]} 条')
    print('完成！')


if __name__ == '__main__':
    main()
