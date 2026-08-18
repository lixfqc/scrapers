# -*- coding: utf-8 -*-
"""
JADA 日本汽车销量爬虫
数据源: https://www.jada.or.jp/
数据类型: 品牌别登录台数、燃料别登录台数
"""
import sys
sys.path.insert(0, '.')
import os
import re
import io
import logging
import requests
import psycopg2
from psycopg2.extras import RealDictCursor
import openpyxl
from datetime import datetime, date
from kba_crawler import BaseCrawler, DB_CONFIG

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s | %(levelname)s | %(message)s'
)

# ============================================
# 日语品牌映射表
# ============================================
JAPANESE_BRAND_MAP = {
    'ダイハツ': 'DAIHATSU',
    'ホンダ': 'HONDA',
    'レクサス': 'LEXUS',
    'マツダ': 'MAZDA',
    '三菱': 'MITSUBISHI',
    '日産': 'NISSAN',
    'スバル': 'SUBARU',
    'スズキ': 'SUZUKI',
    'トヨタ': 'TOYOTA',
    '輸入車': 'IMPORT',
    'ヤマハ': 'YAMAHA',
    'スズカ': 'SUZUKI',
    '本田': 'HONDA',
    '三菱自動車': 'MITSUBISHI',
    '鈴木': 'SUZUKI',
    '豊田': 'TOYOTA',
    '富士重工': 'SUBARU',
    '八王子': 'HACHIOJI',
    'その他': 'OTHER',
    '計': 'TOTAL',
    '合計': 'TOTAL',
    # 商用车品牌
    'いすゞ': 'ISUZU',
    '日野': 'HINO',
    '三菱ふそう': 'MITSUBISHI FUSO',
    'UDトラックス': 'UD TRUCKS',
    # 英文品牌名直接映射
    'SUBARU': 'SUBARU',
    'ＳＵＢＡＲＵ': 'SUBARU',
    'TOYOTA': 'TOYOTA',
    'HONDA': 'HONDA',
    'NISSAN': 'NISSAN',
    'MAZDA': 'MAZDA',
    'SUZUKI': 'SUZUKI',
    'MITSUBISHI': 'MITSUBISHI',
    'DAIHATSU': 'DAIHATSU',
    'LEXUS': 'LEXUS',
    'HINO': 'HINO',
    'ISUZU': 'ISUZU',
}

# 能源类型映射
ENERGY_TYPE_MAP = {
    'ガソリン': 'GASOLINE',
    'ＨＶ': 'HEV',
    'ＨＥＶ': 'HEV',
    'ＰＨＶ': 'PHEV',
    'ＰＨＥＶ': 'PHEV',
    'ディーゼル': 'DIESEL',
    'Ｄ': 'DIESEL',
    'ＥＶ': 'BEV',
    'ＢＥＶ': 'BEV',
    'ＦＣＶ': 'FCV',
    'その他(*)': 'OTHER',
    'その他': 'OTHER',
}

# JADA Excel下载URL模板（通过relays/download中转）
# 注意：这些URL会随时间更新，实际使用时需要从JADA网站解析最新URL
JADA_BRAND_URLS = {
    2025: 'https://www.jada.or.jp/relays/download/364/1700/2127//?file=/files/libs/6636//202602021027028901.xlsx',
    2024: 'https://www.jada.or.jp/relays/download/364/1700/1913/6636/?file=/files/libs/5188//202502031125273606.xlsx',
    2023: 'https://www.jada.or.jp/relays/download/364/1700/1416/5188/?file=/files/libs/3123//202403251320265877.xlsx',
    2022: 'https://www.jada.or.jp/relays/download/364/1700/1417/3123/?file=/files/libs/3124//202403251320452830.xlsx',
}

JADA_FUEL_URLS = {
    2025: 'https://www.jada.or.jp/relays/download/342/1584/2145//?file=/files/libs/6964//202604021716241683.xlsx',
    2024: 'https://www.jada.or.jp/relays/download/342/1584/1920/6964/?file=/files/libs/5211//20250204155758310.xlsx',
    2023: 'https://www.jada.or.jp/relays/download/342/1584/1118/5211/?file=/files/libs/3173//202403251352543040.xlsx',
    2022: 'https://www.jada.or.jp/relays/download/342/1584/1119/3173/?file=/files/libs/3174//202403251353157142.xlsx',
}

# 当前年份实时数据URL（含当年所有月份）
JADA_BRAND_CURRENT_URL = 'https://www.jada.or.jp/files/libs/7425//2026080311130245.xlsx'
JADA_FUEL_CURRENT_URL = 'https://www.jada.or.jp/files/libs/7434//202608041348315964.xlsx'


def _map_brand_name(brand_raw):
    """映射日语品牌名到英文
    
    Args:
        brand_raw: 原始品牌名（日语）
    
    Returns:
        英文品牌名，如果未找到映射则返回原值
    """
    brand_clean = str(brand_raw).strip()
    
    # 直接映射
    if brand_clean in JAPANESE_BRAND_MAP:
        return JAPANESE_BRAND_MAP[brand_clean]
    
    # 模糊匹配（去除特殊字符）
    for jp_name, en_name in JAPANESE_BRAND_MAP.items():
        if brand_clean == jp_name:
            return en_name
    
    return brand_clean


def _map_energy_type(energy_raw):
    """映射日语能源类型到标准能源类型
    
    Args:
        energy_raw: 原始能源类型（日语）
    
    Returns:
        标准能源类型，如果未找到映射则返回原值
    """
    energy_clean = str(energy_raw).strip().upper()
    
    if energy_clean in ENERGY_TYPE_MAP:
        return ENERGY_TYPE_MAP[energy_clean]
    
    return energy_clean


class JadaCrawler(BaseCrawler):
    """JADA日本汽车销量爬虫"""
    
    def __init__(self):
        super().__init__(source_name='jada', country_code='JP')
        self.headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
        }
    
    def download_brand_excel(self, year):
        """下载指定年份的品牌别Excel文件
        
        Args:
            year: 年份
        
        Returns:
            bytes: Excel文件内容，如果下载失败返回None
        """
        self.logger.info(f'下载 {year} 年品牌别Excel...')
        
        if year == datetime.now().year:
            url = JADA_BRAND_CURRENT_URL
        elif year in JADA_BRAND_URLS:
            url = JADA_BRAND_URLS[year]
        else:
            self.logger.warning(f'无 {year} 年的品牌别数据URL')
            return None
        
        try:
            resp = requests.get(url, headers=self.headers, timeout=30)
            if resp.status_code == 200:
                self.logger.info(f'下载成功: {len(resp.content)} bytes')
                return resp.content
            else:
                self.logger.error(f'下载失败: HTTP {resp.status_code}')
                return None
        except Exception as e:
            self.logger.error(f'下载异常: {e}')
            return None
    
    def download_fuel_excel(self, year):
        """下载指定年份的燃料别Excel文件
        
        Args:
            year: 年份
        
        Returns:
            bytes: Excel文件内容，如果下载失败返回None
        """
        self.logger.info(f'下载 {year} 年燃料别Excel...')
        
        if year == datetime.now().year:
            url = JADA_FUEL_CURRENT_URL
        elif year in JADA_FUEL_URLS:
            url = JADA_FUEL_URLS[year]
        else:
            self.logger.warning(f'无 {year} 年的燃料别数据URL')
            return None
        
        try:
            resp = requests.get(url, headers=self.headers, timeout=30)
            if resp.status_code == 200:
                self.logger.info(f'下载成功: {len(resp.content)} bytes')
                return resp.content
            else:
                self.logger.error(f'下载失败: HTTP {resp.status_code}')
                return None
        except Exception as e:
            self.logger.error(f'下载异常: {e}')
            return None
    
    def parse_brand_excel(self, excel_content):
        """解析品牌别Excel文件
        
        Excel结构:
        - 每个Sheet代表一个月份（如"2025年12月"）
        - 表头在第5行
        - 品牌数据行：列2为"合計"，列5为乘用车当月销量，列9为商用车当月销量
        - 比率行：列2为"前年比"，需跳过
        
        Args:
            excel_content: Excel文件内容(bytes)
        
        Returns:
            list: 品牌数据记录列表
        """
        wb = openpyxl.load_workbook(io.BytesIO(excel_content))
        records = []
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # 从Sheet名解析年月
            match = re.match(r'(\d+)年(\d+)月', sheet_name)
            if not match:
                self.logger.warning(f'无法解析Sheet名: {sheet_name}')
                continue
            
            year = int(match.group(1))
            month = int(match.group(2))
            
            self.logger.info(f'解析 {year}-{month:02d} 品牌数据...')
            
            # 从第6行开始遍历（跳过表头行）
            for row in ws.iter_rows(min_row=6, values_only=True):
                if not row or len(row) < 14:
                    continue
                
                brand_raw = row[1]  # 列1: 品牌名
                row_type = row[2]   # 列2: 行类型（"合計"或"前年比"）
                
                # 跳过比率行和空行
                if not brand_raw or not row_type:
                    continue
                if row_type == '前年比':
                    continue
                
                # 只处理"合計"行
                if row_type != '合計':
                    continue
                
                brand_name = _map_brand_name(brand_raw)
                
                # 乘用车当月销量（列5: 計）
                passenger_sales = row[5] if len(row) > 5 else 0
                # 商用车当月销量（列9: 計）
                commercial_sales = row[9] if len(row) > 9 else 0
                
                # 转换为整数
                try:
                    passenger_sales = int(passenger_sales) if passenger_sales else 0
                except (ValueError, TypeError):
                    passenger_sales = 0
                
                try:
                    commercial_sales = int(commercial_sales) if commercial_sales else 0
                except (ValueError, TypeError):
                    commercial_sales = 0
                
                # 跳过TOTAL行（汇总数据）
                if brand_name in ('TOTAL', 'OTHER'):
                    continue
                
                # 添加乘用车记录
                if passenger_sales > 0:
                    records.append({
                        'country_code': 'JP',
                        'source_month': date(year, month, 1),
                        'brand_name_raw': brand_raw,
                        'brand_id': None,  # 将在save_sales时匹配
                        'model_name': None,
                        'vehicle_type': 'passenger',
                        'energy_type': None,
                        'segment': None,
                        'raw_unit': 'units',
                        'sales_volume_raw': passenger_sales,
                        'sales_volume_normalized': passenger_sales,
                        'revision_no': 1,
                        'is_latest': True,
                        'pub_date': None,
                        'crawl_time': datetime.now(),
                        'data_source': 'jada',
                        'notes': 'JADA品牌别数据',
                    })
                
                # 添加商用车记录
                if commercial_sales > 0:
                    records.append({
                        'country_code': 'JP',
                        'source_month': date(year, month, 1),
                        'brand_name_raw': brand_raw,
                        'brand_id': None,
                        'model_name': None,
                        'vehicle_type': 'commercial',
                        'energy_type': None,
                        'segment': None,
                        'raw_unit': 'units',
                        'sales_volume_raw': commercial_sales,
                        'sales_volume_normalized': commercial_sales,
                        'revision_no': 1,
                        'is_latest': True,
                        'pub_date': None,
                        'crawl_time': datetime.now(),
                        'data_source': 'jada',
                        'notes': 'JADA品牌别数据',
                    })
        
        self.logger.info(f'解析到 {len(records)} 条品牌数据记录')
        return records
    
    def parse_fuel_excel(self, excel_content):
        """解析燃料别Excel文件
        
        Excel结构:
        - 每个Sheet代表一个月份（如"2025年12月"）
        - 能源类型表头在第6行
        - 品牌数据行：列2为空，列3/5/7/9/11/13/15为各能源类型销量
        - 构成比行：列2为"構成比"，需跳过
        
        Args:
            excel_content: Excel文件内容(bytes)
        
        Returns:
            list: 能源类型数据记录列表
        """
        wb = openpyxl.load_workbook(io.BytesIO(excel_content))
        records = []
        
        # 能源类型列映射（列索引 -> 能源类型）
        fuel_columns = [
            (3, 'GASOLINE'),   # ガソリン
            (5, 'HEV'),       # ＨＶ
            (7, 'PHEV'),      # ＰＨＶ
            (9, 'DIESEL'),    # ディーゼル
            (11, 'BEV'),      # ＥＶ
            (13, 'FCV'),      # ＦＣＶ
            (15, 'OTHER'),    # その他(*)
        ]
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # 从Sheet名解析年月
            match = re.match(r'(\d+)年(\d+)月', sheet_name)
            if not match:
                self.logger.warning(f'无法解析Sheet名: {sheet_name}')
                continue
            
            year = int(match.group(1))
            month = int(match.group(2))
            
            self.logger.info(f'解析 {year}-{month:02d} 燃料数据...')
            
            # 从第9行开始遍历（跳过表头和比率行）
            for row in ws.iter_rows(min_row=9, values_only=True):
                if not row or len(row) < 18:
                    continue
                
                brand_raw = row[1]  # 列1: 品牌名
                row_type = row[2]   # 列2: 行类型（"構成比"或空）
                
                # 跳过构成比行和空行
                if not brand_raw:
                    continue
                if row_type == '構成比':
                    continue
                
                brand_name = _map_brand_name(brand_raw)
                
                # 跳过TOTAL行
                if brand_name in ('TOTAL', 'OTHER'):
                    continue
                
                # 提取各能源类型销量
                for col_idx, energy_type in fuel_columns:
                    if col_idx < len(row):
                        sales = row[col_idx]
                        if sales is None or sales == '':
                            continue
                        
                        try:
                            sales = int(sales)
                        except (ValueError, TypeError):
                            continue
                        
                        if sales > 0:
                            records.append({
                                'country_code': 'JP',
                                'source_month': date(year, month, 1),
                                'brand_name_raw': brand_raw,
                                'brand_id': None,
                                'model_name': None,
                                'vehicle_type': 'passenger',
                                'energy_type': energy_type,
                                'segment': None,
                                'raw_unit': 'units',
                                'sales_volume_raw': sales,
                                'sales_volume_normalized': sales,
                                'revision_no': 1,
                                'is_latest': True,
                                'pub_date': None,
                                'crawl_time': datetime.now(),
                                'data_source': 'jada',
                                'notes': 'JADA燃料别数据',
                            })
        
        self.logger.info(f'解析到 {len(records)} 条能源类型数据记录')
        return records
    
    def crawl_year(self, year, data_type='brand'):
        """爬取指定年份的数据
        
        Args:
            year: 年份
            data_type: 'brand'(品牌别) 或 'fuel'(燃料别)
        
        Returns:
            int: 保存的记录数
        """
        if data_type == 'brand':
            excel_content = self.download_brand_excel(year)
            if not excel_content:
                self.logger.error(f'{year} 年品牌别数据下载失败')
                return 0
            
            records = self.parse_brand_excel(excel_content)
        elif data_type == 'fuel':
            excel_content = self.download_fuel_excel(year)
            if not excel_content:
                self.logger.error(f'{year} 年燃料别数据下载失败')
                return 0
            
            records = self.parse_fuel_excel(excel_content)
        else:
            self.logger.error(f'未知数据类型: {data_type}')
            return 0
        
        if not records:
            self.logger.warning(f'{year} 年无数据记录')
            return 0
        
        # 保存记录
        saved_count = 0
        for record in records:
            if self.save_sales(record):
                saved_count += 1
        
        self.logger.info(f'{year} 年保存 {saved_count} 条 {data_type} 记录')
        return saved_count
    
    def crawl_range(self, start_year, end_year, data_type='brand'):
        """爬取指定年份范围的数据
        
        Args:
            start_year: 开始年份
            end_year: 结束年份
            data_type: 'brand'(品牌别) 或 'fuel'(燃料别)
        
        Returns:
            dict: 各年份保存的记录数
        """
        results = {}
        
        for year in range(start_year, end_year + 1):
            self.logger.info(f'=== 爬取 {year} 年 {data_type} 数据 ===')
            count = self.crawl_year(year, data_type)
            results[year] = count
        
        return results
    
    def get_brand_id(self, brand_name_raw):
        """获取品牌ID（重写基类方法，支持日语品牌名映射）
        
        Args:
            brand_name_raw: 原始品牌名（日语）
        
        Returns:
            int or None: 品牌ID
        """
        brand_name = _map_brand_name(brand_name_raw)
        
        try:
            # 使用独立连接，避免干扰主连接
            conn = psycopg2.connect(**DB_CONFIG)
            cur = conn.cursor(cursor_factory=RealDictCursor)
            
            # 优先匹配canonical_name
            cur.execute("""
                SELECT id FROM brand_name_mapping 
                WHERE LOWER(canonical_name) = LOWER(%s) 
                  AND status = 'active'
                LIMIT 1
            """, (brand_name,))
            row = cur.fetchone()
            if row:
                cur.close()
                conn.close()
                return row['id']
            
            # 匹配brand_name_cn
            cur.execute("""
                SELECT id FROM brand_name_mapping 
                WHERE LOWER(brand_name_cn) = LOWER(%s) 
                  AND status = 'active'
                LIMIT 1
            """, (brand_name,))
            row = cur.fetchone()
            if row:
                cur.close()
                conn.close()
                return row['id']
            
            cur.close()
            conn.close()
            
        except Exception as e:
            self.logger.error(f'品牌ID查询失败: {e}')
        
        return None
    
    def save_sales(self, record):
        """重写保存方法，在保存前自动获取品牌ID"""
        record['brand_id'] = self.get_brand_id(record['brand_name_raw'])
        if record['brand_id']:
            self.logger.debug(f'品牌匹配成功: {record["brand_name_raw"]} -> {record["brand_id"]}')
        return super().save_sales(record)


def main():
    """主函数：爬取日本JADA数据"""
    crawler = JadaCrawler()
    
    print('=== JADA日本汽车销量爬虫 ===')
    print('\n可用数据URL:')
    print('品牌别: 2022-2025年 + 2026年至今')
    print('燃料别: 2022-2025年 + 2026年至今')
    
    # 爬取品牌别数据（2024-2026年）
    print('\n--- 爬取品牌别数据 ---')
    brand_results = crawler.crawl_range(2024, 2026, 'brand')
    for year, count in brand_results.items():
        print(f'  {year}年: {count} 条记录')
    
    # 爬取燃料别数据（2024-2026年）
    print('\n--- 爬取燃料别数据 ---')
    fuel_results = crawler.crawl_range(2024, 2026, 'fuel')
    for year, count in fuel_results.items():
        print(f'  {year}年: {count} 条记录')
    
    print('\n完成！')


if __name__ == '__main__':
    main()
